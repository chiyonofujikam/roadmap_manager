"""
Helper functions for roadmap management operations.

This module provides utility functions for:
    - XML export/import
    - Reading collaborator lists from Excel
    - Building Excel interfaces with data validation
    - CLI argument parsing
    - Logging configuration

Author: Mustapha ELKAMILI
"""
import argparse
import io
import logging
import os
import shutil
import stat
import sys
import tempfile
import time
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import (DataValidation,
                                               DataValidationList)

# xlwings is imported lazily only when needed (in add_validation_list function)
# This avoids slow startup when xlwings is not required

def get_exe_dir() -> Path:
    """
    Determine the directory where the executable/script is located.

    Returns:
        Path: The directory where the executable/script is located.
    """
    # sys.argv[0] contains the path to the script/executable that was invoked
    # This works for both .exe launchers and regular Python scripts
    script_path = Path(sys.argv[0]).resolve()

    # Check if sys.argv[0] points to an .exe file
    if script_path.suffix.lower() == '.exe' and script_path.exists():
        # Running as an .exe file (even if it's a launcher script), Use the directory where the .exe file is located
        exe_dir = script_path.parent
    elif getattr(sys, 'frozen', False):
        # Running as a frozen/packaged executable
        exe_dir = Path(sys.executable).parent.resolve()
    else:
        # Running as a Python script - use the script's directory
        # If __file__ is available, use it; otherwise fall back to script's directory
        try:
            exe_dir = Path(__file__).parent.parent.resolve()  # Go up from helpers.py to roadmap/ to project root
        except NameError:
            exe_dir = script_path.parent

    # Create .logs directory in the same directory as the executable
    logs_dir = exe_dir / ".logs"
    try:
        logs_dir.mkdir(parents=True, exist_ok=True)
        # Verify the directory was created
        if not logs_dir.exists():
            raise OSError(f"Failed to create logs directory: {logs_dir}")
    except (OSError, PermissionError):
        # If we can't create in exe directory, fall back to current working directory
        logs_dir = Path.cwd() / ".logs"
        try:
            logs_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            # Last resort: use temp directory
            logs_dir = Path(tempfile.gettempdir()) / "roadmap_logs"
            logs_dir.mkdir(parents=True, exist_ok=True)

    return str(logs_dir / "roadmap.log")

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(get_exe_dir(), mode="a", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def zip_folder(folder_path: Path, zip_path: Path) -> None:
    """
    Create a zip archive of a folder.

    Args:
        folder_path (Path): Path to the folder to zip.
        zip_path (Path): Path where the zip file should be created.

    Returns:
        None

    Note:
        Creates a zip file containing all files and subdirectories from the source folder.
        Preserves the folder structure within the zip archive, including the folder name itself.
    """
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_path in folder_path.rglob('*'):
            if not file_path.is_file():
                continue
            # Preserve folder structure: include folder name in zip
            # e.g., if folder_path is "RM_Collaborateurs", zip will contain "RM_Collaborateurs/file.xlsx"
            arcname = file_path.relative_to(folder_path.parent)
            zipf.write(file_path, arcname)

def rmtree_with_retry(folder_path: Path, max_retries: int = 5) -> bool:
    """
    Remove a directory tree with retry logic for Windows/OneDrive locks.

    Args:
        folder_path (Path): Path to the folder to remove.
        max_retries (int): Maximum number of retry attempts.

    Returns:
        bool: True if removal succeeded, False otherwise.
    """
    def on_rm_error(func, path, exc_info):
        """Error handler for shutil.rmtree to handle read-only files."""
        try:
            os.chmod(path, stat.S_IWRITE)
            func(path)
        except Exception:
            pass

    for attempt in range(max_retries):
        try:
            shutil.rmtree(folder_path, onerror=on_rm_error)
            return True
        except PermissionError as e:
            if attempt < max_retries - 1:
                logger.warning(f"[RMTREE] Attempt {attempt + 1}/{max_retries} failed: {e}. Retrying in 2s...")
                time.sleep(2)
            else:
                logger.error(f"[RMTREE] Failed to remove folder after {max_retries} attempts: {e}")
                return False
    return False

def write_xml(rows: list, xml_output: Path) -> None:
    """
    Write data rows to XML file in format expected by VBA.

    Creates an XML file with a structure that VBA can easily parse. Each row
    becomes a <row> element containing <col1>, <col2>, etc. child elements.

    Args:
        rows (list): List of row data, where each row is a list of values.
            None values are converted to empty strings.
        xml_output (Path): Path where the XML file should be written.

    Returns:
        None

    Example:
        >>> rows = [["Alice", 100], ["Bob", 200]]
        >>> write_xml(rows, Path("output.xml"))
        Creates XML with two <row> elements.
    """
    root = ET.Element("rows")

    for r in rows:
        row_el = ET.SubElement(root, "row")
        for idx, val in enumerate(r, start=1):
            col = ET.SubElement(row_el, f"col{idx}")
            col.text = "" if val is None else str(val)

    tree = ET.ElementTree(root)
    tree.write(xml_output, encoding="utf-8", xml_declaration=True)


def add_data_validations_to_sheet(ws_pointage, start_row: int = 3) -> None:
    """
    Add standard data validation lists to POINTAGE sheet.

    Creates data validation lists for columns D (week), E (key), F (label), and G (function).
    This function centralizes the validation creation logic used across multiple methods.

    Args:
        ws_pointage: openpyxl worksheet object for the POINTAGE sheet.
        start_row (int, optional): Starting row for validation ranges. Defaults to 3.

    Returns:
        None
    """
    # Clear existing validations to avoid duplicates
    if hasattr(ws_pointage, 'data_validations') and ws_pointage.data_validations is not None:
        if hasattr(ws_pointage.data_validations, 'dataValidation'):
            ws_pointage.data_validations.dataValidation = []
    else:
        ws_pointage.data_validations = DataValidationList()

    # Create standard data validation lists
    dv_semaine = DataValidation(type="list", formula1="='POINTAGE'!$A$2:$A$2")
    dv_cle = DataValidation(type="list", formula1="='LC'!$B$3:$B$10000")
    dv_libelle = DataValidation(type="list", formula1="='LC'!$C$3:$C$10000")
    dv_fonction = DataValidation(type="list", formula1="='LC'!$D$3:$D$10000")

    validations = [
        (dv_semaine, 'D'),
        (dv_cle, 'E'),
        (dv_libelle, 'F'),
        (dv_fonction, 'G'),
    ]

    for dv, col in validations:
        ws_pointage.add_data_validation(dv)
        dv.ranges.add(f"{col}{start_row}:{col}1000")

def get_collaborators(synthese_file: Path | str) -> list[str]:
    """
    Extract collaborator names from XML file.

    Reads from collabs.xml file in the base directory to avoid permission errors.
    The XML file is deleted after reading.

    Args:
        synthese_file (Path | str): Path to the synthesis Excel file (used as base
            directory to locate collabs.xml).

    Returns:
        list[str]: List of collaborator names, stripped of whitespace.
        Returns empty list if XML file cannot be read or doesn't exist.

    Note:
        The collabs.xml file will be deleted after reading.
    """
    collabs = []
    synthese_file = Path(synthese_file)

    # Check if collabs.xml exists in the same directory as synthese_file
    xml_file = synthese_file.parent / "collabs.xml"

    if not xml_file.exists():
        logger.info(f"[GET_COLLABORATORS] collabs.xml file not found: {xml_file}")
        return collabs

    try:
        logger.info(f"[GET_COLLABORATORS] Reading from XML file: {xml_file}")
        tree = ET.parse(xml_file)
        root = tree.getroot()

        for collab_elem in root.findall("collaborator"):
            collab_name = collab_elem.text
            if collab_name and collab_name.strip():
                collabs.append(collab_name.strip())

        logger.info(f"[GET_COLLABORATORS] Read {len(collabs)} collaborators from XML")

        # Delete the XML file after reading
        try:
            xml_file.unlink()
            logger.info(f"[GET_COLLABORATORS] Deleted XML file: {xml_file}")
        except Exception as del_err:
            logger.warning(f"[GET_COLLABORATORS] Could not delete XML file: {del_err}")

        return collabs
    except Exception as xml_err:
        logger.error(f"[GET_COLLABORATORS] Error reading XML file: {xml_err}")

    return collabs

def load_lc_excel(base_dir: Path | str) -> list[list]:
    """
    Load LC (conditional lists) data from LC.xlsx file.

    Reads LC data from LC.xlsx file in the base directory.
    Reads columns B-I (columns 2-9) from the LC sheet, starting at row 2.

    Args:
        base_dir (Path | str): Base directory path where LC.xlsx should be located.

    Returns:
        list[list]: List of row data, where each row is a list of 8 values (columns B-I).
        Empty strings are converted to None for Excel compatibility.
        Returns empty list if Excel file cannot be read or doesn't exist.

    Note:
        The LC.xlsx file will be deleted after reading.
    """
    lc_data = []
    base_dir = Path(base_dir)
    excel_file = base_dir / "LC.xlsx"

    if not excel_file.exists():
        logger.warning(f"[LOAD_LC_EXCEL] LC.xlsx file not found: {excel_file}")
        return lc_data

    try:
        logger.info(f"[LOAD_LC_EXCEL] Reading from Excel file: {excel_file}")
        # Use data_only=False to access cell formatting information
        wb = load_workbook(excel_file, data_only=False, read_only=True)

        if "LC" not in wb.sheetnames:
            logger.warning(f"[LOAD_LC_EXCEL] LC sheet not found in {excel_file}")
            wb.close()
            return lc_data

        ws = wb["LC"]

        # Use iter_rows for bulk reading (much faster than cell-by-cell)
        # Read columns B-I (columns 2-9), starting at row 2
        # Limit to first 10000 rows for safety
        max_row = max(ws.max_row, 10000)

        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=9, values_only=False):
            # Check if row has any non-empty data
            has_data = False
            row_data = []

            for cell in row:
                if cell.value is not None:
                    from datetime import datetime, date
                    
                    # Priority 1: If cell is explicitly formatted as text (@), treat as string
                    # Even if openpyxl reads it as a date, we want the text representation
                    if cell.number_format == '@':
                        # Cell is formatted as text - read as string
                        # If it's a date object, convert to string preserving format
                        if isinstance(cell.value, (datetime, date)):
                            # For dates in text-formatted cells, we need to preserve the display format
                            # Since VBA already converted to text, this shouldn't happen, but handle it
                            # Use the cell's internal value converted to string
                            cell_str = str(cell.value)
                        else:
                            cell_str = str(cell.value)
                    elif isinstance(cell.value, (datetime, date)):
                        # Cell contains a date/datetime object but is NOT formatted as text
                        # This shouldn't happen if VBA worked correctly, but handle it
                        # Format as ISO date to ensure consistency
                        if isinstance(cell.value, datetime):
                            cell_str = cell.value.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            cell_str = cell.value.strftime('%Y-%m-%d')
                    else:
                        # For non-date values, convert to string normally
                        cell_str = str(cell.value)
                    
                    cell_str = cell_str.strip()
                    if cell_str:
                        row_data.append(cell_str)
                        has_data = True
                    else:
                        row_data.append(None)
                else:
                    row_data.append(None)

            # Stop if row is empty (assuming data is contiguous)
            if not has_data:
                break

            # Add row to data
            lc_data.append(row_data)

        wb.close()
        logger.info(f"[LOAD_LC_EXCEL] Loaded {len(lc_data)} rows of LC data from Excel")

        # Delete the Excel file after reading
        try:
            excel_file.unlink()
            logger.info(f"[LOAD_LC_EXCEL] Deleted Excel file: {excel_file}")
        except Exception as del_err:
            logger.warning(f"[LOAD_LC_EXCEL] Could not delete Excel file: {del_err}")

        return lc_data
    except Exception as excel_err:
        logger.error(f"[LOAD_LC_EXCEL] Error reading Excel file: {excel_err}")

    return lc_data

def build_interface(template_bytes: bytes, output_path: str, collab_name: str) -> None:
    """
    Build a single collaborator interface Excel file from template.

    Creates a new Excel file for a collaborator based on the template.
    Sets the collaborator name in cell B1 and adds data validation lists
    for pointage entry (week, key, label, function).

    Args:
        template_bytes (bytes): Binary content of the template Excel file.
        output_path (str): Path where the new interface file should be saved.
        collab_name (str): Name of the collaborator to set in the interface.

    Returns:
        None

    Note:
        This function is designed to be called in parallel processes.
        It uses bytes instead of file path to avoid file locking issues in parallel execution.

    Data Validation Lists:
        - Column D: Week (from POINTAGE!A2:A2)
        - Column E: Key (from LC!B3:B1000)
        - Column F: Label (from LC!C3:C1000)
        - Column G: Function (from LC!D3:D1000)
    """
    wb = load_workbook(filename=io.BytesIO(template_bytes))
    ws_pointage = wb["POINTAGE"]

    # Write collaborator name
    ws_pointage["B1"].value = collab_name

    # Add data validations (using row 3 to match other methods)
    add_data_validations_to_sheet(ws_pointage, start_row=3)

    wb.save(output_path)
    wb.close()

def add_validation_list(
    wb,
    list_range: str,
    target_column: str,
    dropdown_sheet: str = "POINTAGE",
    list_sheet: str = "LC",
    end_row: int = 1000,
    start_row: int = 4):
    """
    Add data validation list to Excel workbook using xlwings.

    Applies a dropdown list validation to a column range in Excel.
    The dropdown options come from a specified range in another sheet.
    
    xlwings is imported lazily here to avoid startup delay when not needed.

    Args:
        wb: xlwings Book object representing the Excel workbook.
        list_range (str): Excel range string (e.g., "B3:B1000") containing
            the list of valid values.
        target_column (str): Column letter (e.g., "E") where validation
            should be applied.
        dropdown_sheet (str, optional): Name of sheet containing target cells.
            Defaults to "POINTAGE".
        list_sheet (str, optional): Name of sheet containing source list.
            Defaults to "LC".
        end_row (int, optional): Last row number for validation range.
            Defaults to 1000.
        start_row (int, optional): First row number for validation range.
            Defaults to 4.

    Returns:
        Book: The modified xlwings Book object.

    Note:
        Removes any existing validation on the target range before adding new validation.
        Uses Excel's native Validation API via xlwings.
    """
    import xlwings as xw  # Lazy import to speed up startup
    # Build source address for formula
    ws_list = wb.sheets[list_sheet]
    source_range = ws_list.range(list_range)

    # Build target range: column from start_row to end of Excel
    ws_dropdown = wb.sheets[dropdown_sheet]
    target_range = ws_dropdown.range(f"{target_column}{start_row}:{target_column}{end_row}")

    # Apply validation
    validation = target_range.api.Validation
    try:
        validation.Delete()
    except Exception:
        pass

    validation.Add(
        Type=3,
        AlertStyle=1,
        Operator=1,
        Formula1=f"='{list_sheet}'!{source_range.get_address()}"
    )

    return wb

def get_parser() -> argparse.ArgumentParser:
    """
    Create and configure CLI argument parser.

    Sets up argument parser with subcommands for create, delete, pointage, and update operations.
    Each subcommand has its own specific options.

    Returns:
        argparse.ArgumentParser: Configured argument parser ready to parse command-line arguments.

    Commands:
        - create: Create collaborator interfaces
            Options: --way (normal/para/xlw), --archive
        - delete: Delete collaborator interfaces
            Options: --archive, --force
        - pointage: Export time tracking data
            Options: --delete
        - update: Update conditional lists

    Global Options:
        --basedir: Base directory for file operations
    """
    parser = argparse.ArgumentParser(
        description="Roadmap Management CLI - Automate CE VHST roadmap operations including time tracking, interface creation, and data synchronization.",
        usage="roadmap <command>",
    )

    # Global argument
    parser.add_argument(
        "--basedir",
        type=str,
        default="none",
        help="Specify the base directory path containing roadmap files. If not provided, uses platform-specific default or current directory."
    )

    subparsers_action = parser.add_subparsers(dest="action", required=True)
    create_parser = subparsers_action.add_parser("create", help="Generate Excel interface files for all collaborators listed in the synthesis file")
    create_parser.add_argument(
        "--way",
        choices=['xlw', 'normal', 'para'],
        default='normal',
        help="Processing mode: 'normal' for sequential processing (~50s), 'para' for parallel processing (~9s, fastest), or 'xlw' for xlwings-based processing (~3min, best for VBA integration)"
    )

    delete_parser = subparsers_action.add_parser("delete", help="Remove or archive all collaborator interface files from RM_Collaborateurs folder")
    delete_parser.add_argument(
        "--archive",
        action="store_true",
        help="Copy files to Archived folder before moving to Deleted folder. Without this flag, files are moved directly to Deleted folder"
    )

    delete_parser.add_argument(
        "--force",
        action="store_true",
        help="Required flag to confirm deletion operation. Without this flag, the operation will be aborted with a warning"
    )

    subparsers_action.add_parser("pointage", help="Export time tracking data from collaborator Excel files to XML format for VBA import")
    subparsers_action.add_parser("update", help="Synchronize conditional lists (LC) from master synthesis file to template and all collaborator interface files")
    subparsers_action.add_parser("cleanup", help="Delete interface files for collaborators that are missing from the XML list")

    return parser
