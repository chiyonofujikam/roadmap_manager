"""
Main script for CE VHST Roadmap automation.

This module provides the RoadmapManager class and CLI interface for automating roadmap management tasks including:
    1. Pointage (time tracking export)
    2. Updating conditional lists (LC)
    3. Creating user interfaces
    4. Deleting interfaces

The module integrates with Excel files using openpyxl and xlwings, and can be called from both command-line and VBA macros.

Author: Mustapha EL KAMILI
"""
import platform
import shutil
import sys
import tempfile
import time
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime
from pathlib import Path

import xlwings as xw
from openpyxl import load_workbook
from tqdm import tqdm

from roadmap.helpers import (add_data_validations_to_sheet, add_validation_list, app,
                              build_interface, get_collaborators, get_parser, logger,
                              write_xml, zip_folder, rmtree_with_retry)


class RoadmapManager:
    """
    Manages roadmap interfaces and data for CE VHST collaborators.

    This class handles the creation, deletion, and management of Excel-based roadmap interfaces.
    It coordinates between the master synthesis file, template files, and individual collaborator files.

    Attributes:
        base_path (Path): Base directory path for all roadmap files.
        synthese_file (Path): Path to the master synthesis Excel file.
        template_file (Path): Path to the template Excel file.
        rm_folder (Path): Directory containing collaborator interface files.
        archived_folder (Path): Directory for archived files.
        deleted_folder (Path): Directory for deleted files.
        xml_output (Path): Path for pointage XML export file.
        all_ok (bool): Flag indicating if all required files exist.

    Example:
        >>> manager = RoadmapManager(base_dir="/path/to/roadmap")
        >>> manager.create_interfaces(archive=True)
        >>> manager.pointage()
    """

    def __init__(self, base_dir: str | Path):
        """
        Initialize RoadmapManager with base directory.

        Sets up all necessary paths and validates that required files exist.
        Creates necessary directories if they don't exist.

        Args:
            base_dir (str | Path): Base directory path containing roadmap files.
                Should contain 'Synthèse_RM_CE.xlsm' and 'RM_template.xlsx'.

        Note:
            Logs an error if required files are missing, but continues initialization.
            Check `all_ok` attribute before operations.
        """
        self.base_path = Path(base_dir)

        # Define paths for configuration files
        self.synthese_file = self.base_path / "Synthèse_RM_CE.xlsm"
        self.template_file = self.base_path / "RM_template.xlsx"

        # Define and create necessary folders
        self.rm_folder = self.base_path / "RM_Collaborateurs"
        self.archived_folder = self.base_path / "Archived"
        self.deleted_folder = self.base_path / "Deleted"
        self.xml_output = self.base_path / "pointage_output.xml"

        for folder in [self.rm_folder, self.archived_folder, self.deleted_folder]:
            folder.mkdir(exist_ok=True)

        # Check existence of essential files
        self.all_ok = all([
            self.synthese_file.exists(),
            self.template_file.exists()
        ])

        if not self.all_ok:
            logger.error("Required files 'Synthese_RM_CE.xlsm' or 'RM_template.xlsx' are missing. Please check the base directory.")

    def check_roadmap_archive(self) -> Path:
        """
        Archive existing 'RM_Collaborateurs' folder if it contains files.

        Zips the existing 'RM_Collaborateurs' folder and saves it to the Archived directory with a timestamped name.
        If the folder doesn't exist or is empty, creates a new empty folder.

        Returns:
            Path: Path to the archived zip file (if archived) or the new
                'RM_Collaborateurs' folder (if empty/non-existent).

        Note:
            Only archives if there are .xlsx files in the folder.
        """
        if self.rm_folder.exists():
            rm_count = sum(1 for f in self.rm_folder.glob("*.xlsx") if f.is_file())
            zip_path = self.archived_folder / f"Archive_RM_Collaborateurs_{datetime.now():%d%m%Y_%H%M%S}.zip"
            if rm_count != 0:
                zip_folder(self.rm_folder, zip_path)

                # Remove the original folder after zipping (with retry for OneDrive/Windows locks)
                if not rmtree_with_retry(self.rm_folder):
                    logger.warning("[CHECK_ROADMAP_ARCHIVE] Could not remove original folder, but archive was created")
                else:
                    logger.info(f"[CHECK_ROADMAP_ARCHIVE] Archived {rm_count} interface file(s) to {zip_path.name}")
                
                # Recreate the folder for new interfaces
                self.rm_folder.mkdir(exist_ok=True)
                return zip_path

        self.rm_folder.mkdir(exist_ok=True)
        return self.rm_folder

    def create_interfaces_fast(self, archive: bool, max_workers: int = 8) -> None:
        """
            Create user interfaces using parallel processing with openpyxl.

            This method uses ProcessPoolExecutor to create interfaces in parallel, significantly faster than sequential processing.
            Estimated time: ~9s for 51 files (vs ~50s sequential).

            Args:
                archive (bool): If True, archives existing RM_Collaborateurs folder
                    before creating new interfaces.
                max_workers (int, optional): Maximum number of parallel workers.
                    Defaults to 8.

            Returns:
                None: Returns early if required files are missing or no collaborators found.

            Note:
                Requires template file to be closed. Logs error if template is locked by another process.
        """
        if not self.all_ok:
            return

        logger.info("[CREATE_INTERFACES] Parallel processing mode interface creation")

        collaborators = get_collaborators(self.synthese_file)
        if not collaborators:
            logger.info(
                "[CREATE_INTERFACES] the list of CE is empty."
                f" Please check 'Gestion_Interfaces' sheet in '{self.synthese_file}'")
            return

        logger.info(f"[CREATE_INTERFACES] Found {len(collaborators)} collaborators")

        if archive:
            self.check_roadmap_archive()

        try:
            template_bytes = Path(self.template_file).read_bytes()
        except PermissionError:
            logger.error(f"'{self.template_file}' is opened. Please close the excel file")
            return

        futures = []
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            for collab in collaborators:
                output_path = str(self.rm_folder / f"RM_{collab}.xlsx")
                futures.append(
                    executor.submit(build_interface, template_bytes, output_path, collab)
                )

            for future in tqdm(futures, desc="Creating interfaces (parallel)"):
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"error: {e}")

        logger.info("[CREATE_INTERFACES] parallel creation complete.")

    def create_interfaces(self, archive: bool) -> None:
        """
        Create user interfaces using sequential processing with openpyxl.

        Creates individual Excel interface files for each collaborator listed in the 'Synthese_RM_CE.xlsm' file.
        Each interface is based on the template file and includes data validation lists for pointage entry.

        Args:
            archive (bool): If True, archives existing 'RM_Collaborateurs' folder before creating new interfaces.

        Returns:
            None: Returns early if required files are missing or no collaborators found.

        Note:
            Estimated time: ~50s for 51 files. For faster processing, use 'create_interfaces_fast()' instead.
        """
        if not self.all_ok:
            return

        logger.info("[CREATE_INTERFACES] interface creation (Normal processing mode)")

        collaborators = get_collaborators(self.synthese_file)
        if not collaborators:
            logger.info(
                "[CREATE_INTERFACES] the list of CE is empty."
                f" Please check 'Gestion_Interfaces' sheet in '{self.synthese_file}'")
            return

        logger.info(f"[CREATE_INTERFACES] Found {len(collaborators)} collaborators, synthese file : {self.synthese_file.name}")

        if archive:
            self.check_roadmap_archive()

        for collab in tqdm(collaborators, desc="Creating interfaces", total=len(collaborators)):
            target = self.rm_folder / f"RM_{collab}.xlsx"

            try:
                wb = load_workbook(self.template_file)
            except PermissionError:
                logger.error(f"'{self.template_file}' is opened. Please close the excel file")
                return

            ws_pointage = wb["POINTAGE"]

            # Write collaborator name
            ws_pointage["B1"].value = collab

            # Add data validations
            add_data_validations_to_sheet(ws_pointage, start_row=3)

            wb.save(target)
            wb.close()

        logger.info("[CREATE_INTERFACES] creation done.")

    def create_interfaces_xlwings(self, archive: bool) -> None:
        """
        Create user interfaces using xlwings library.

        Uses xlwings for Excel automation, which provides better integration with Excel's native features but is slower than openpyxl.
        Useful when you need Excel to be running or for VBA integration scenarios.

        Args:
            archive (bool): If True, archives existing 'RM_Collaborateurs' folder before creating new interfaces.

        Returns:
            None: Returns early if required files are missing or no collaborators found.

        Note:
            Estimated time: ~3min4s for 51 files. This is the slowest method.
                Prefer 'create_interfaces()' or 'create_interfaces_fast()' unless xlwings-specific features are required.
        """
        if not self.all_ok:
            return

        logger.info("[CREATE_INTERFACES] interface creation (xlwings processing mode)")

        collaborators = get_collaborators(self.synthese_file)
        if not collaborators:
            logger.info(
                "[CREATE_INTERFACES] the list of CE is empty."
                f" Please check 'Gestion_Interfaces' sheet in '{self.synthese_file}'")
            return

        logger.info(f"[CREATE_INTERFACES] Found {len(collaborators)} collaborators")

        if archive:
            self.check_roadmap_archive()

        for collab in tqdm(collaborators, desc="Creating interfaces", total=len(collaborators)):
            target_path = self.rm_folder / f"RM_{collab}.xlsx"
            shutil.copy2(self.template_file, target_path)

            wb = xw.Book(str(target_path))
            pointage_sheet = wb.sheets["POINTAGE"]
            pointage_sheet["B1"].value = collab

            # Semain
            wb = add_validation_list(
                wb,
                list_sheet="POINTAGE",
                list_range="A2:A2",
                target_column="D",
            )

            # Clef d'imputation
            wb = add_validation_list(
                wb,
                list_range="B3:B1000",
                target_column="E",
            )

            # Libellé
            wb = add_validation_list(
                wb,
                list_range="C3:C1000",
                target_column="F",
            )

            # Fonction
            wb = add_validation_list(
                wb,
                list_range="D3:D1000",
                target_column="G",
            )

            wb.save()
            wb.close()

        logger.info("[CREATE_INTERFACES] creation done.")

    def delete_interfaces(self, archive: bool) -> None:
        """
        Delete or archive all collaborator interface files.

        Zips the 'RM_Collaborateurs' folder and saves it to the 'Deleted' directory with a timestamped name.
        Optionally archives the folder first before deletion.

        Args:
            archive (bool): If True, archives the folder to 'Archived' directory before moving to 'Deleted' directory.
                If False, moves directly to 'Deleted' folder.

        Returns:
            None: Returns early if required files are missing or folder doesn't exist.

        Note:
            This operation is destructive. Files are zipped and saved to 'Deleted' folder, preserving originals if archive=True.
        """
        if not self.all_ok:
            return

        logger.info("[DELETE_INTERFACES] Starting interface deletion")

        rm_folder = self.rm_folder
        if not rm_folder.exists():
            logger.warning("[DELETE_INTERFACES] RM_Collaborateurs folder does not exist")
            return

        rm_count = sum(1 for f in rm_folder.glob("*.xlsx") if f.is_file())

        if archive:
            archived_path = self.check_roadmap_archive()
            logger.info(f"[DELETE_INTERFACES] Archived {rm_count} interface file(s) to {archived_path.name}")

            # Copy the archived zip to Deleted folder (if it's a zip file)
            try:
                target_zip = self.deleted_folder / f"Deleted_RM_Collaborateurs_{datetime.now().strftime('%d%m%Y_%H%M%S')}.zip"
                if rm_count != 0 and archived_path.exists() and archived_path.suffix == '.zip':
                    shutil.copy2(archived_path, target_zip)
                    logger.info(f"[DELETE_INTERFACES] Deleted & Moved {rm_count} interface file(s) to {target_zip.name}")
                    return
            except Exception as e:
                logger.error(f"[DELETE_INTERFACES] Error while copying zip file: {e}")
                return

        try:
            target_zip = self.deleted_folder / f"Deleted_RM_Collaborateurs_{datetime.now().strftime('%d%m%Y_%H%M%S')}.zip"
            if rm_count != 0:
                zip_folder(rm_folder, target_zip)

                # Remove the original folder after zipping
                shutil.rmtree(rm_folder)
                logger.info(f"[DELETE_INTERFACES] Deleted & Moved {rm_count} interface file(s) to {target_zip.name}")
                return
        except Exception as e:
            logger.error(f"[DELETE_INTERFACES] Error while zipping folder: {e}")
            return

    def pointage(self) -> bool:
        """
        Export pointage (time tracking) data from collaborator files to XML.

        Reads time tracking data from all collaborator Excel files in the 'RM_Collaborateurs' folder and exports it to a single XML file.
        The XML format is designed to be consumed by VBA macros in Excel.

        Reads data from the 'POINTAGE' sheet, starting at row 4, columns A-K.
        Stops reading when encountering a fully empty row.

        Returns:
            bool: True if data was exported, False if no data found or operation failed. Always creates XML file (empty if no data).

        Note:
            Creates an empty XML file if no data exists, as VBA expects the file to be present. Skips temporary Excel files (starting with '~$').
        """
        if not self.all_ok:
            return False

        if not self.rm_folder.exists():
            logger.error("RM_Collaborateurs folder not found")
            return False

        collaborator_files = [
            filepath for filepath in self.rm_folder.glob("*.xlsx")
            if not filepath.name.startswith("~$")
        ]

        if not collaborator_files:
            logger.warning("No collaborator files found")
            write_xml([], self.xml_output)
            return False

        logger.info(f"[POINTAGE] Processing {len(collaborator_files)} collaborator files")

        all_rows = []

        for collaborator_file in collaborator_files:
            logger.info(f"[POINTAGE] Reading {collaborator_file}")

            wb = load_workbook(collaborator_file, data_only=True, read_only=True)
            sheet = wb["POINTAGE"]

            for row in sheet.iter_rows(min_row=4, min_col=1, max_col=11):
                row_data = [cell.value for cell in row]

                # Stop when hitting a fully empty row
                if all(v is None for v in row_data):
                    break

                all_rows.append(row_data)

            wb.close()

        if not all_rows:
            logger.info("[POINTAGE] No data to export → creating EMPTY XML")
            write_xml([], self.xml_output)
            return False

        write_xml(all_rows, self.xml_output)
        logger.info(f"[POINTAGE] XML successfully created with {len(all_rows)} rows → {self.xml_output}")

        return True

    def update_lc(self) -> None:
        """
        Update conditional lists (LC) in 'RM_template.xlsx' and all collaborator interface files.

        Reads LC data from the 'Synthese_RM_CE.xlsm' file and updates the 'LC' sheet in the 'RM_template.xlsx' file and all collaborator interface files.
        This ensures all files have synchronized dropdown list options for data entry.

        Reads LC data from columns B-I (columns 2-9), starting at row 2. Updates cell values while preserving formatting.

        Returns:
            None: Returns early if required files are missing.

        Note:
            Updates both the template file and all files in 'RM_Collaborateurs' folder. Skips temporary Excel files (starting with '~$').
        """
        if not self.all_ok:
            logger.error("[UPDATE_LC] Required files are missing. Cannot proceed.")
            return

        logger.info("[UPDATE_LC] Starting LC update process")

        # Read LC data from SYNTHESE file
        # Try openpyxl first (faster, doesn't require Excel to be running)
        # Fall back to xlwings only if file is open
        lc_data = []
        source_wb = None
        opened_source = False

        try:
            # First try openpyxl (faster and more reliable)
            try:
                logger.info("[UPDATE_LC] Attempting to read LC data using openpyxl...")
                synthese_wb = load_workbook(self.synthese_file, read_only=True, data_only=True)
                lc_sheet = synthese_wb["LC"]

                # Read data from columns B-I (columns 2-9), starting at row 2
                for row in lc_sheet.iter_rows(min_row=2, min_col=2, max_col=9):
                    row_data = [cell.value for cell in row]

                    # Stop if all cells in the row are None/empty
                    if all(cell is None for cell in row_data):
                        break

                    lc_data.append(row_data)

                synthese_wb.close()
                logger.info(f"[UPDATE_LC] Loaded {len(lc_data)} rows of LC data using openpyxl")

            except PermissionError:
                # File is open - use xlwings to read from it
                logger.info("[UPDATE_LC] File is open, attempting to read using xlwings...")
                try:
                    # Try to connect if already open (with timeout handling)
                    try:
                        source_wb = xw.Book(str(self.synthese_file))
                        opened_source = False
                        logger.info("[UPDATE_LC] Connected to open Excel file")
                    except Exception as connect_err:
                        # If connection fails, try opening it
                        logger.info(f"[UPDATE_LC] Connection failed, opening file: {connect_err}")
                        source_wb = app.books.open(str(self.synthese_file))
                        opened_source = True

                    # Read LC data directly from the open file using xlwings
                    lc_sheet = source_wb.sheets["LC"]

                    # Read data from columns B-I (columns 2-9), starting at row 2
                    row = 2
                    while True:
                        # Read a row from columns B to I (2 to 9)
                        row_data = []
                        for col in range(2, 10):  # Columns B-I
                            try:
                                cell_value = lc_sheet.range((row, col)).value
                                row_data.append(cell_value)
                            except Exception as cell_err:
                                logger.warning(f"[UPDATE_LC] Error reading cell ({row}, {col}): {cell_err}")
                                row_data.append(None)

                        # Stop if all cells in the row are None/empty
                        if all(cell is None for cell in row_data):
                            break

                        lc_data.append(row_data)
                        row += 1

                        # Safety limit to prevent infinite loops
                        if row > 10000:
                            logger.warning("[UPDATE_LC] Reached safety limit of 10000 rows, stopping")
                            break

                    logger.info(f"[UPDATE_LC] Loaded {len(lc_data)} rows of LC data using xlwings")

                except Exception as xlw_err:
                    logger.error(f"[UPDATE_LC] Error reading SYNTHESE file with xlwings: {xlw_err}")
                    raise

            except Exception as e:
                logger.error(f"[UPDATE_LC] Error reading SYNTHESE file: {e}")
                raise

        except Exception as e:
            logger.error(f"[UPDATE_LC] Failed to read LC data from SYNTHESE file: {e}. Please ensure the SYNTHESE file exists and is accessible")
            return
        finally:
            # Only close if we opened it (don't close if user has it open)
            if opened_source and source_wb:
                try:
                    source_wb.close()
                except Exception as close_err:
                    logger.warning(f"[UPDATE_LC] Error closing workbook: {close_err}")

        if not lc_data:
            logger.warning("[UPDATE_LC] No LC data found. Nothing to update.")
            return

        # Update template file
        logger.info("[UPDATE_LC] Updating template file...")
        try:
            self._update_lc_in_file(self.template_file, lc_data)
        except Exception as e:
            logger.error(f"[UPDATE_LC] Error updating template file: {e}")

        # Update all collaborator files
        if self.rm_folder.exists():
            rm_files = list(self.rm_folder.glob("*.xlsx"))
            rm_files = [f for f in rm_files if not f.name.startswith("~$")]
            logger.info(f"[UPDATE_LC] Updating {len(rm_files)} collaborator files...")

            for rm_file in rm_files:
                try:
                    self._update_lc_in_file(rm_file, lc_data)
                except Exception as e:
                    logger.error(f"[UPDATE_LC] Error updating {rm_file.name}: {e}")
                    # Continue with other files even if one fails

        logger.info("[UPDATE_LC] LC update completed")

    def _update_lc_in_file(self, file_path: Path, lc_data: list) -> None:
        """
        Update 'LC' sheet in a single Excel file.

        Private helper method that updates the 'LC' (conditional lists) sheet in a given Excel file with new data from the 'Synthese_RM_CE.xlsm' file.

        Args:
            file_path (Path): Path to the Excel file to update.
            lc_data (list): List of row data to write to the 'LC' sheet. Each row is a list of cell values.

        Returns:
            None: Logs warning and returns early if 'LC' sheet not found.

        Note:
            Updates cell values only, preserving all formatting (colors, borders, fonts, etc.).
            For collaborator files (RM_*.xlsx), recreates data validation lists in POINTAGE sheet
            exactly as they would be created in create_interfaces(), ensuring consistency.
            Existing data in the file is preserved - only the LC sheet and data validation are updated.
            Clears cells beyond the new data range to remove old data.
            Uses xlwings to handle files that may be open in Excel.
        """
        # Create temporary file for working copy
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_path = Path(tmp.name)

        wb = None

        try:
            # Try to copy the file using shutil (works if file is not open)
            # If file is open, this will fail - user needs to save and close it first
            try:
                shutil.copy2(file_path, temp_path)
            except PermissionError:
                logger.warning(f"[UPDATE_LC] Cannot update {file_path.name} - it may be open in Excel. Skipping this file.")
                return
            except Exception as e:
                logger.warning(f"[UPDATE_LC] Error copying {file_path.name}: {e}. Skipping this file.")
                return

            # Load workbook from temp file - use data_only=False to preserve formulas and data validation
            wb = load_workbook(temp_path, data_only=False)

            if "LC" not in wb.sheetnames:
                logger.warning(f"[UPDATE_LC] LC sheet not found in {file_path.name}")
                wb.close()
                return

            lc_sheet = wb["LC"]

            # Update cells with new data (preserves formatting by only updating values)
            for row_idx, row in enumerate(lc_sheet.iter_rows(min_row=2, min_col=2, max_col=9), start=2):
                if row_idx - 2 >= len(lc_data):
                    for cell in row:
                        if cell.column >= 2:
                            cell.value = None
                    continue

                row_data = lc_data[row_idx - 2]
                for cell in row:
                    col_idx = cell.column - 2
                    if col_idx < len(row_data):
                        cell.value = row_data[col_idx]
                    else:
                        cell.value = None

            # Recreate data validation lists for collaborator files (same as create_interfaces())
            # This ensures consistent validation regardless of file state
            is_collab_file = file_path.name.startswith("RM_") or "RM_Collaborateurs" in str(file_path.parent)

            if is_collab_file and "POINTAGE" in wb.sheetnames:
                ws_pointage = wb["POINTAGE"]
                try:
                    logger.info(f"[UPDATE_LC] Recreating data validation lists for collaborator file {file_path.name}")
                    add_data_validations_to_sheet(ws_pointage, start_row=3)
                    logger.info(f"[UPDATE_LC] Successfully recreated data validation lists in {file_path.name}")
                except Exception as e:
                    logger.error(f"[UPDATE_LC] Error recreating data validation in {file_path.name}: {e}")

            # Save modified workbook to temp file
            wb.save(temp_path)
            wb.close()
            wb = None

            # Small delay to ensure file handle is released
            time.sleep(0.1)

            # Copy modified temp file back to original location
            # This works even if the original file is open (we overwrite it)
            shutil.copy2(temp_path, file_path)

        except Exception as e:
            logger.error(f"[UPDATE_LC] Error updating LC in {file_path.name}: {e}")
            # Don't raise - allow other files to be processed
            return
        finally:
            # Clean up temporary file
            if temp_path.exists():
                max_retries = 5
                for attempt in range(max_retries):
                    try:
                        temp_path.unlink()
                        break
                    except (PermissionError, OSError):
                        if attempt < max_retries - 1:
                            time.sleep(0.2)
                        else:
                            logger.warning(f"[UPDATE_LC] Could not delete temporary file {temp_path}, but operation completed successfully")


def main() -> None:
    """
    Main entry point for CLI interface.

    Parses command-line arguments and executes the appropriate RoadmapManager operation.
    Supports 'create', 'delete', 'pointage', and 'update' commands.

    Default base directory is platform-specific:
        - Windows: 'C:\\Users\\MustaphaELKAMILI\\OneDrive - IKOSCONSULTING\\test_RM\\files'
        - Other: '/mnt/c/Users/MustaphaELKAMILI/OneDrive - IKOSCONSULTING/test_RM/files'

    Can be overridden with '--basedir' argument.

    Returns:
        None
    """
    if platform.system() == "Windows":
        BASE_DIR = Path(r"C:\Users\MustaphaELKAMILI\OneDrive - IKOSCONSULTING\test_RM\files")
    else:
        BASE_DIR = Path("/mnt/c/Users/MustaphaELKAMILI/OneDrive - IKOSCONSULTING/test_RM/files")

    args = get_parser().parse_args()
    manager = RoadmapManager(base_dir=BASE_DIR if args.basedir == 'none' else Path(args.basedir))

    if args.action == "create":
        if args.way == 'normal':
            manager.create_interfaces(args.archive)
        elif args.way == 'para':
            manager.create_interfaces_fast(args.archive)
        elif args.way == 'xlw':
            manager.create_interfaces_xlwings(args.archive)
        else:
            logger.error(f"Unknown '--way' argument '{args.way}'. Valid choices are 'normal', 'para', and 'xlw'.")
        return

    if args.action == "delete":
        if not args.force:
            logger.warning("⚠️  Operation not confirmed. Use --force to proceed.")
            return

        manager.delete_interfaces(archive=args.archive)
        return

    if args.action == "pointage":
        manager.pointage()
        return

    if args.action == "update":
        try:
            manager.update_lc()
        except Exception as e:
            logger.error(f"Fatal error in update_lc: {e}", exc_info=True)
            sys.exit(1)


def run() -> None:
    """
    Entry point for console script installation.

    This function is called when the 'roadmap' command is invoked from the command line.
    It's registered as a console script in 'pyproject.toml'.

    Returns:
        None
    """
    main()


if __name__ == "__main__":
    main()
