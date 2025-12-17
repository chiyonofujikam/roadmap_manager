"""
RoadmapManager class for CE VHST Roadmap automation.

This module provides the RoadmapManager class for automating roadmap management tasks including:
    1. Pointage (time tracking export)
    2. Updating conditional lists (LC)
    3. Creating user interfaces
    4. Deleting interfaces

The class integrates with Excel files using openpyxl, and can be called from both command-line and VBA macros.

Author: Mustapha EL KAMILI
"""
import shutil
import tempfile
import time
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from tqdm import tqdm

from roadmap.helpers import (add_data_validations_to_sheet, build_interface,
                             get_collaborators, load_lc_excel, logger,
                             rmtree_with_retry, write_xml, zip_folder)


class RoadmapManager:
    """
    Manages roadmap interfaces and data for CE VHST collaborators.

    This class handles the creation, deletion, and management of Excel-based roadmap interfaces.
    It coordinates between the master synthesis file, template files, and individual collaborator files.

    Attributes:
        base_path (Path): Base directory path for all roadmap files.
        synthese_file (Path): Path to the master synthesis Excel file.
        template_file (Path): Path to the template Excel file.
        rm_folder (Path): Directory containing collaborator interface files.
        archived_folder (Path): Directory for archived files.
        deleted_folder (Path): Directory for deleted files.
        xml_output (Path): Path for pointage XML export file.
        all_ok (bool): Flag indicating if all required files exist.

    Example:
        >>> manager = RoadmapManager(base_dir="/path/to/roadmap")
        >>> manager.create_interfaces(archive=True)
        >>> manager.pointage()
    """

    def __init__(self, base_dir: str | Path):
        """
        Initialize RoadmapManager with base directory.

        Sets up all necessary paths and validates that required files exist.
        Creates necessary directories if they don't exist.

        Args:
            base_dir (str | Path): Base directory path containing roadmap files.
                Should contain 'Synthèse_RM_CE.xlsm' and 'RM_template.xlsx'.

        Note:
            Logs an error if required files are missing, but continues initialization.
            Check `all_ok` attribute before operations.
        """
        self.base_path = Path(base_dir)

        # Define paths for configuration files
        self.synthese_file = self.base_path / "Synthèse_RM_CE.xlsm"
        self.template_file = self.base_path / "RM_template.xlsx"

        # Define and create necessary folders
        self.rm_folder = self.base_path / "RM_Collaborateurs"
        self.archived_folder = self.base_path / "Archived"
        self.deleted_folder = self.base_path / "Deleted"
        self.xml_output = self.base_path / "pointage_output.xml"

        for folder in [self.rm_folder, self.archived_folder, self.deleted_folder]:
            folder.mkdir(exist_ok=True)

        # Check existence of essential files
        self.all_ok = all([
            self.synthese_file.exists(),
            self.template_file.exists()
        ])

        if not self.all_ok:
            logger.error("Required files 'Synthese_RM_CE.xlsm' or 'RM_template.xlsx' are missing. Please check the base directory.")

    def create_interfaces_fast(self, max_workers: int = 8) -> None:
        """
        Create user interfaces using parallel processing with openpyxl.

        This method uses ProcessPoolExecutor to create interfaces in parallel, significantly faster than sequential processing.
        Only creates missing collaborator files - checks if file exists before creating.

        Args:
            archive (bool): Ignored - kept for backward compatibility.
            max_workers (int, optional): Maximum number of parallel workers.
                Defaults to 8.

        Returns:
            None: Returns early if required files are missing or no collaborators found.

        Note:
            Requires template file to be closed. Logs error if template is locked by another process.
            Only creates files that don't already exist.
        """
        if not self.all_ok:
            return

        logger.info("[CREATE_INTERFACES] Parallel processing mode interface creation")

        collaborators = get_collaborators(self.synthese_file)
        if not collaborators:
            logger.info(
                "[CREATE_INTERFACES] the list of CE is empty."
                f" Please check XML file or 'Gestion_Interfaces' sheet in '{self.synthese_file}'")
            return

        logger.info(f"[CREATE_INTERFACES] Found {len(collaborators)} collaborators")

        # Ensure RM_Collaborateurs folder exists
        self.rm_folder.mkdir(exist_ok=True)

        try:
            template_bytes = Path(self.template_file).read_bytes()
        except PermissionError:
            logger.error(f"'{self.template_file}' is opened. Please close the excel file")
            return

        # Filter out collaborators whose files already exist
        missing_collabs = []
        for collab in collaborators:
            output_path = self.rm_folder / f"RM_{collab}.xlsx"
            if not output_path.exists():
                missing_collabs.append(collab)
            else:
                logger.debug(f"[CREATE_INTERFACES] File already exists: {output_path.name}")

        if not missing_collabs:
            logger.info("[CREATE_INTERFACES] All collaborator files already exist. Nothing to create.")
            return

        logger.info(f"[CREATE_INTERFACES] Creating {len(missing_collabs)} missing interface file(s)")

        futures = []
        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            for collab in missing_collabs:
                output_path = str(self.rm_folder / f"RM_{collab}.xlsx")
                futures.append(
                    executor.submit(build_interface, template_bytes, output_path, collab)
                )

            for future in tqdm(futures, desc="Creating interfaces (parallel)"):
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"error: {e}")

        logger.info("[CREATE_INTERFACES] parallel creation complete.")

    def create_interfaces(self) -> None:
        """
        Create user interfaces using sequential processing with openpyxl.

        Creates individual Excel interface files for each collaborator from XML file.
        Only creates missing collaborator files - checks if file exists before creating.
        Each interface is based on the template file and includes data validation lists for pointage entry.

        Returns:
            None: Returns early if required files are missing or no collaborators found.

        Note:
            Estimated time: ~50s for 51 files. For faster processing, use 'create_interfaces_fast()' instead.
            Only creates files that don't already exist.
        """
        if not self.all_ok:
            return

        logger.info("[CREATE_INTERFACES] interface creation (Normal processing mode)")

        collaborators = get_collaborators(self.synthese_file)
        if not collaborators:
            logger.info(
                "[CREATE_INTERFACES] the list of CE is empty."
                f" Please check XML file or 'Gestion_Interfaces' sheet in '{self.synthese_file}'")
            return

        logger.info(f"[CREATE_INTERFACES] Found {len(collaborators)} collaborators, synthese file : {self.synthese_file.name}")

        # Ensure RM_Collaborateurs folder exists
        self.rm_folder.mkdir(exist_ok=True)

        # Filter out collaborators whose files already exist
        missing_collabs = []
        for collab in collaborators:
            target = self.rm_folder / f"RM_{collab}.xlsx"
            if not target.exists():
                missing_collabs.append(collab)
            else:
                logger.debug(f"[CREATE_INTERFACES] File already exists: {target.name}")

        if not missing_collabs:
            logger.info("[CREATE_INTERFACES] All collaborator files already exist. Nothing to create.")
            return

        logger.info(f"[CREATE_INTERFACES] Creating {len(missing_collabs)} missing interface file(s)")

        for collab in tqdm(missing_collabs, desc="Creating interfaces", total=len(missing_collabs)):
            target = self.rm_folder / f"RM_{collab}.xlsx"

            try:
                wb = load_workbook(self.template_file)
            except PermissionError:
                logger.error(f"'{self.template_file}' is opened. Please close the excel file")
                return

            ws_pointage = wb["POINTAGE"]

            # Write collaborator name
            ws_pointage["B1"].value = collab

            # Add data validations
            add_data_validations_to_sheet(ws_pointage, start_row=3)

            wb.save(target)
            wb.close()

        logger.info("[CREATE_INTERFACES] creation done.")

    def delete_and_archive_interfaces(self, archive: bool) -> None:
        """
        Delete or archive the entire RM_Collaborateurs folder.

        Archives the 'RM_Collaborateurs' folder to 'Archived' directory if archive=True,
        then zips and moves the folder to 'Deleted' directory with a timestamped name.
        The entire folder is removed after archiving/deleting.

        Args:
            archive (bool): If True, archives the folder to 'Archived' directory before moving to 'Deleted' directory.
                If False, moves directly to 'Deleted' folder.

        Returns:
            None: Returns early if required files are missing or folder doesn't exist.

        Note:
            This operation is destructive. The entire RM_Collaborateurs folder is removed.
        """
        if not self.all_ok:
            return

        logger.info("[DELETE_INTERFACES] Starting interface deletion")

        rm_folder = self.rm_folder
        if not rm_folder.exists():
            logger.warning("[DELETE_INTERFACES] RM_Collaborateurs folder does not exist")
            return

        rm_count = sum(1 for f in rm_folder.glob("*.xlsx") if f.is_file())

        if rm_count == 0:
            logger.warning("[DELETE_INTERFACES] RM_Collaborateurs folder is empty")
            # Still delete the folder if it exists
            try:
                rmtree_with_retry(rm_folder)
                logger.info("[DELETE_INTERFACES] Empty folder removed")
            except Exception as e:
                logger.error(f"[DELETE_INTERFACES] Error removing empty folder: {e}")
            return

        timestamp = datetime.now().strftime('%d%m%Y_%H%M%S')

        # Archive to Archived folder if requested
        if archive:
            try:
                archived_zip = self.archived_folder / f"Archive_RM_Collaborateurs_{timestamp}.zip"
                zip_folder(rm_folder, archived_zip)
                logger.info(f"[DELETE_INTERFACES] Archived {rm_count} interface file(s) to {archived_zip.name}")
            except Exception as e:
                logger.error(f"[DELETE_INTERFACES] Error while archiving folder: {e}")
                return

        # Move to Deleted folder
        try:
            deleted_zip = self.deleted_folder / f"Deleted_RM_Collaborateurs_{timestamp}.zip"
            zip_folder(rm_folder, deleted_zip)

            # Remove the original folder after zipping
            if not rmtree_with_retry(rm_folder):
                logger.warning("[DELETE_INTERFACES] Could not remove original folder, but zip was created")
            else:
                logger.info(f"[DELETE_INTERFACES] Deleted & Moved {rm_count} interface file(s) to {deleted_zip.name}")
        except Exception as e:
            logger.error(f"[DELETE_INTERFACES] Error while zipping folder: {e}")
            return

    def delete_missing_collaborators(self) -> None:
        """
        Delete interface files for collaborators that are missing from the XML list.

        Compares existing files in RM_Collaborateurs folder with the collaborator list from XML.
        If a file exists but the collaborator is not in the XML list, that file is deleted.

        Returns:
            None: Returns early if required files are missing or folder doesn't exist.

        Note:
            Only deletes files that don't match any collaborator in the current XML list.
            Skips temporary Excel files (starting with '~$').
        """
        if not self.all_ok:
            return

        logger.info("[DELETE_MISSING_COLLABORATORS] Starting cleanup of missing collaborators")

        if not self.rm_folder.exists():
            logger.warning("[DELETE_MISSING_COLLABORATORS] RM_Collaborateurs folder does not exist")
            return

        # Get current list of collaborators from XML
        collaborators = get_collaborators(self.synthese_file)
        if not collaborators:
            logger.warning("[DELETE_MISSING_COLLABORATORS] No collaborators found in XML. Skipping cleanup.")
            return

        # Create a set of expected file names for faster lookup
        expected_files = {f"RM_{collab}.xlsx" for collab in collaborators}
        logger.info(f"[DELETE_MISSING_COLLABORATORS] Found {len(collaborators)} collaborators in XML")

        # Get all existing files in the folder
        existing_files = [
            f for f in self.rm_folder.glob("*.xlsx")
            if f.is_file() and not f.name.startswith("~$")
        ]

        if not existing_files:
            logger.info("[DELETE_MISSING_COLLABORATORS] No files found in RM_Collaborateurs folder")
            return

        # Find files that don't match any collaborator in the list
        files_to_delete = []
        for file_path in existing_files:
            if file_path.name not in expected_files:
                files_to_delete.append(file_path)

        if not files_to_delete:
            logger.info("[DELETE_MISSING_COLLABORATORS] All files match collaborators in XML. Nothing to delete.")
            return

        logger.info(f"[DELETE_MISSING_COLLABORATORS] Found {len(files_to_delete)} file(s) to delete")

        # Create zip archive of files to delete before deletion
        timestamp = datetime.now().strftime('%d%m%Y_%H%M%S')
        zip_filename = f"Deleted_Missing_RM_collaborators_{timestamp}.zip"
        zip_path = self.deleted_folder / zip_filename

        # Create temporary folder to hold files for zipping
        temp_folder = None
        try:
            temp_folder = tempfile.mkdtemp(prefix="missing_collabs_", dir=self.base_path)
            temp_folder_path = Path(temp_folder)

            # Copy files to temporary folder
            for file_path in files_to_delete:
                dest_path = temp_folder_path / file_path.name
                shutil.copy2(file_path, dest_path)
                logger.debug(f"[DELETE_MISSING_COLLABORATORS] Copied to temp folder: {file_path.name}")

            # Zip the temporary folder using zip_folder function
            zip_folder(temp_folder_path, zip_path)
            logger.info(f"[DELETE_MISSING_COLLABORATORS] Created archive: {zip_filename}")
        except Exception as e:
            logger.error(f"[DELETE_MISSING_COLLABORATORS] Error creating zip archive: {e}")
            # Continue with deletion even if zip fails
        finally:
            # Clean up temporary folder
            if temp_folder and Path(temp_folder).exists():
                try:
                    shutil.rmtree(temp_folder)
                    logger.debug(f"[DELETE_MISSING_COLLABORATORS] Cleaned up temporary folder")
                except Exception as e:
                    logger.warning(f"[DELETE_MISSING_COLLABORATORS] Could not remove temporary folder: {e}")

        # Delete the orphaned files
        deleted_count = 0
        for file_path in files_to_delete:
            try:
                file_path.unlink()
                logger.info(f"[DELETE_MISSING_COLLABORATORS] Deleted: {file_path.name}")
                deleted_count += 1
            except PermissionError:
                logger.warning(f"[DELETE_MISSING_COLLABORATORS] Cannot delete {file_path.name} - file may be open in Excel")
            except Exception as e:
                logger.error(f"[DELETE_MISSING_COLLABORATORS] Error deleting {file_path.name}: {e}")

        logger.info(f"[DELETE_MISSING_COLLABORATORS] Cleanup complete. Deleted {deleted_count} file(s). Archive saved to: {zip_filename}")

    def pointage(self) -> bool:
        """
        Export pointage (time tracking) data from collaborator files to XML.

        Reads time tracking data from all collaborator Excel files in the 'RM_Collaborateurs' folder and exports it to a single XML file.
        The XML format is designed to be consumed by VBA macros in Excel.

        Reads data from the 'POINTAGE' sheet, starting at row 4, columns A-K.
        Stops reading when encountering a fully empty row.

        Returns:
            bool: True if data was exported, False if no data found or operation failed. Always creates XML file (empty if no data).

        Note:
            Creates an empty XML file if no data exists, as VBA expects the file to be present. Skips temporary Excel files (starting with '~$').
        """
        if not self.all_ok:
            return False

        if not self.rm_folder.exists():
            logger.error("RM_Collaborateurs folder not found")
            return False

        collaborator_files = [
            filepath for filepath in self.rm_folder.glob("*.xlsx")
            if not filepath.name.startswith("~$")
        ]

        if not collaborator_files:
            logger.warning("No collaborator files found")
            write_xml([], self.xml_output)
            return False

        logger.info(f"[POINTAGE] Processing {len(collaborator_files)} collaborator files")

        all_rows = []

        for collaborator_file in collaborator_files:
            logger.info(f"[POINTAGE] Reading {collaborator_file}")

            wb = load_workbook(collaborator_file, data_only=True, read_only=True)
            sheet = wb["POINTAGE"]

            k1_value = sheet["K1"].value or 0

            for row in sheet.iter_rows(min_row=4, min_col=1, max_col=11):
                row_data = [cell.value for cell in row]

                # Stop when hitting a fully empty row
                if all(v is None for v in row_data):
                    break

                # Append K1 total to help downstream coloring logic
                row_data.append(k1_value)
                all_rows.append(row_data)

            wb.close()

        if not all_rows:
            logger.info("[POINTAGE] No data to export → creating EMPTY XML")
            write_xml([], self.xml_output)
            return False

        write_xml(all_rows, self.xml_output)
        logger.info(f"[POINTAGE] XML successfully created with {len(all_rows)} rows → {self.xml_output}")

        return True

    def update_lc(self) -> None:
        """
        Update conditional lists (LC) in 'RM_template.xlsx' and all collaborator interface files.

        Reads LC data from the 'LC.xlsx' file (generated by btn_update_LC button) and updates the 'LC' sheet 
        in the 'RM_template.xlsx' file and all collaborator interface files.
        This ensures all files have synchronized dropdown list options for data entry.

        Reads LC data from LC.xlsx file which contains data from columns B-I (columns 2-9), starting at row 2.
        Updates cell values while preserving formatting. Also updates data validation lists for collaborator files.

        Returns:
            None: Returns early if required files are missing or LC.xlsx not found.

        Note:
            Updates both the template file and all files in 'RM_Collaborateurs' folder. 
            Skips temporary Excel files (starting with '~$').
            For collaborator files, recreates data validation lists to ensure consistency with updated LC data.
        """
        if not self.all_ok:
            logger.error("[UPDATE_LC] Required files are missing. Cannot proceed.")
            return

        logger.info("[UPDATE_LC] Starting LC update process")

        # Read LC data from LC.xlsx file (generated by VBA button)
        lc_data = load_lc_excel(self.base_path)

        if not lc_data:
            logger.warning("[UPDATE_LC] No LC data found in LC.xlsx. Nothing to update.")
            return

        logger.info(f"[UPDATE_LC] Loaded {len(lc_data)} rows of LC data from LC.xlsx")

        # Update template file
        logger.info("[UPDATE_LC] Updating template file...")
        try:
            self._update_lc_in_file(self.template_file, lc_data)
        except Exception as e:
            logger.error(f"[UPDATE_LC] Error updating template file: {e}")

        # Update all collaborator files
        if self.rm_folder.exists():
            rm_files = list(self.rm_folder.glob("*.xlsx"))
            rm_files = [f for f in rm_files if not f.name.startswith("~$")]
            logger.info(f"[UPDATE_LC] Updating {len(rm_files)} collaborator files...")

            for rm_file in rm_files:
                try:
                    self._update_lc_in_file(rm_file, lc_data)
                except Exception as e:
                    logger.error(f"[UPDATE_LC] Error updating {rm_file.name}: {e}")

        logger.info("[UPDATE_LC] LC update completed")

    def _update_lc_in_file(self, file_path: Path, lc_data: list) -> None:
        """
        Update 'LC' sheet in a single Excel file.

        Private helper method that updates the 'LC' (conditional lists) sheet in a given Excel file 
        with new data from LC.xml file.

        Args:
            file_path (Path): Path to the Excel file to update.
            lc_data (list): List of row data to write to the 'LC' sheet. Each row is a list of 8 values 
                (columns B-I). None values represent empty cells.

        Returns:
            None: Logs warning and returns early if 'LC' sheet not found.

        Note:
            Updates cell values only, preserving all formatting (colors, borders, fonts, etc.).
            For collaborator files (RM_*.xlsx), recreates data validation lists in POINTAGE sheet
            exactly as they would be created in create_interfaces(), ensuring consistency.
            This is important because collaborator files may already have data, and updating LC requires
            updating the data validation to match the new LC data.
            Existing data in the file is preserved - only the LC sheet and data validation are updated.
            Clears cells beyond the new data range to remove old data.
        """
        # Create temporary file for working copy
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            temp_path = Path(tmp.name)

        wb = None

        try:
            # Try to copy the file using shutil (works if file is not open)
            # If file is open, this will fail - user needs to save and close it first
            try:
                shutil.copy2(file_path, temp_path)
            except PermissionError:
                logger.warning(f"[UPDATE_LC] Cannot update {file_path.name} - it may be open in Excel. Skipping this file.")
                return
            except Exception as e:
                logger.warning(f"[UPDATE_LC] Error copying {file_path.name}: {e}. Skipping this file.")
                return

            # Load workbook from temp file - use data_only=False to preserve formulas and data validation
            wb = load_workbook(temp_path, data_only=False)

            if "LC" not in wb.sheetnames:
                logger.warning(f"[UPDATE_LC] LC sheet not found in {file_path.name}")
                wb.close()
                return

            lc_sheet = wb["LC"]

            # Find the last row with data to clear old data efficiently
            last_data_row = len(lc_data) + 1  # +1 because we start at row 2
            old_max_row = lc_sheet.max_row
            
            # Clear old data in bulk if it exists beyond our new data
            if old_max_row > last_data_row:
                # Delete rows beyond our data range (much faster than clearing cell by cell)
                lc_sheet.delete_rows(last_data_row + 1, old_max_row - last_data_row)

            # Write new data starting at row 2 using bulk operations
            # Prepare data as a list of lists for efficient writing
            # Set all cells to text format to prevent date interpretation
            for row_idx, row_data in enumerate(lc_data, start=2):
                # Write all 8 columns at once (B-I = columns 2-9)
                for col_idx in range(8):
                    excel_col = col_idx + 2  # Excel column number (B=2, C=3, ..., I=9)
                    cell = lc_sheet.cell(row=row_idx, column=excel_col)
                    # Set cell format to text FIRST to prevent date interpretation
                    cell.number_format = '@'  # '@' is Excel's text format code
                    if col_idx < len(row_data) and row_data[col_idx] is not None:
                        # Convert to string and ensure it's written as text
                        # This prevents Excel from interpreting date-like strings as dates
                        value_str = str(row_data[col_idx]).strip()
                        # Write as string value - the '@' format ensures it stays as text
                        # Setting number_format before value ensures Excel treats it as text
                        cell.value = value_str
                    else:
                        cell.value = None

            # Recreate data validation lists for collaborator files (same as create_interfaces())
            # This ensures consistent validation regardless of file state
            is_collab_file = file_path.name.startswith("RM_") or "RM_Collaborateurs" in str(file_path.parent)

            if is_collab_file and "POINTAGE" in wb.sheetnames:
                ws_pointage = wb["POINTAGE"]
                try:
                    logger.info(f"[UPDATE_LC] Recreating data validation lists for collaborator file {file_path.name}")
                    add_data_validations_to_sheet(ws_pointage, start_row=3)
                    logger.info(f"[UPDATE_LC] Successfully recreated data validation lists in {file_path.name}")
                except Exception as e:
                    logger.error(f"[UPDATE_LC] Error recreating data validation in {file_path.name}: {e}")

            # Save modified workbook to temp file
            wb.save(temp_path)
            wb.close()
            wb = None

            # Small delay to ensure file handle is released
            time.sleep(0.1)

            # Copy modified temp file back to original location
            # This works even if the original file is open (we overwrite it)
            shutil.copy2(temp_path, file_path)

        except Exception as e:
            logger.error(f"[UPDATE_LC] Error updating LC in {file_path.name}: {e}")
            # Don't raise - allow other files to be processed
            return
        finally:
            # Clean up temporary file
            if temp_path.exists():
                max_retries = 5
                for attempt in range(max_retries):
                    try:
                        temp_path.unlink()
                        break
                    except (PermissionError, OSError):
                        if attempt < max_retries - 1:
                            time.sleep(0.2)
                        else:
                            logger.warning(f"[UPDATE_LC] Could not delete temporary file {temp_path}, but operation completed successfully")
