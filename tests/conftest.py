"""
Pytest configuration and shared fixtures for Roadmap Manager tests.

This module provides reusable fixtures for setting up test environments
with the required file structure and test data.
"""
import io
from pathlib import Path

import pytest
from openpyxl import Workbook

@pytest.fixture
def setup_test_environment(tmp_path):
    """
    Create complete test environment with all required files.

    Creates:
        - RM_template.xlsx with POINTAGE and LC sheets
        - Synthèse_RM_CE.xlsm with Gestion_Interfaces, SYNTHESE, LC sheets
        - collabs.xml with test collaborators
        - Required folders (RM_Collaborateurs, Archived, Deleted)

    Returns:
        Path: The temporary directory path containing the test environment.
    """
    # Create template file
    template = Workbook()
    ws_pointage = template.active
    ws_pointage.title = "POINTAGE"
    ws_pointage["A2"] = "Week 1"
    ws_pointage["B1"] = ""  # Collaborator name placeholder

    ws_lc = template.create_sheet("LC")
    ws_lc["B3"] = "KEY001"
    ws_lc["C3"] = "Label 1"
    ws_lc["D3"] = "Function 1"
    ws_lc["B4"] = "KEY002"
    ws_lc["C4"] = "Label 2"
    ws_lc["D4"] = "Function 2"

    template.save(tmp_path / "RM_template.xlsx")

    # Create synthese file (save as .xlsm extension but it's actually xlsx format for testing)
    synthese = Workbook()
    ws_gi = synthese.active
    ws_gi.title = "Gestion_Interfaces"
    ws_gi["B3"] = "CLIGNIEZ Yann"
    ws_gi["B4"] = "GANI Karim"
    ws_gi["B5"] = "MOUHOUT Marouane"

    ws_synth = synthese.create_sheet("SYNTHESE")
    ws_synth["A1"] = "Header1"
    ws_synth["B1"] = "Header2"

    ws_lc2 = synthese.create_sheet("LC")
    ws_lc2["B3"] = "KEY001"
    ws_lc2["C3"] = "Label 1"

    synthese.save(tmp_path / "Synthèse_RM_CE.xlsm")

    # Create collabs.xml
    xml_content = """<?xml version="1.0" encoding="UTF-8"?>
    <collaborators>
        <collaborator>CLIGNIEZ Yann</collaborator>
        <collaborator>GANI Karim</collaborator>
        <collaborator>MOUHOUT Marouane</collaborator>
    </collaborators>"""
    (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

    # Create necessary folders
    (tmp_path / "RM_Collaborateurs").mkdir()
    (tmp_path / "Archived").mkdir()
    (tmp_path / "Deleted").mkdir()

    return tmp_path


@pytest.fixture
def setup_test_environment_with_interfaces(setup_test_environment):
    """
    Create test environment with pre-created interface files (no data).

    Extends setup_test_environment with empty interface files for each collaborator.

    Returns:
        Path: The temporary directory path containing the test environment.
    """
    tmp_path = setup_test_environment
    rm_folder = tmp_path / "RM_Collaborateurs"

    # Create interface files without data
    for name in ["CLIGNIEZ Yann", "GANI Karim", "MOUHOUT Marouane"]:
        wb = Workbook()
        ws = wb.active
        ws.title = "POINTAGE"
        ws["B1"] = name
        ws["A2"] = "Week 1"

        ws_lc = wb.create_sheet("LC")
        ws_lc["B3"] = "KEY001"
        ws_lc["C3"] = "Label 1"

        wb.save(rm_folder / f"RM_{name}.xlsx")

    return tmp_path


@pytest.fixture
def setup_test_environment_with_data(setup_test_environment):
    """
    Create test environment with populated interface files.

    Extends setup_test_environment with interface files containing pointage data.

    Returns:
        Path: The temporary directory path containing the test environment.
    """
    tmp_path = setup_test_environment
    rm_folder = tmp_path / "RM_Collaborateurs"

    # Create interface files with data
    for idx, name in enumerate(["CLIGNIEZ Yann", "GANI Karim", "MOUHOUT Marouane"], start=1):
        wb = Workbook()
        ws = wb.active
        ws.title = "POINTAGE"
        ws["B1"] = name
        ws["A2"] = "Week 1"

        # Add pointage data (starting at row 4 as per the application)
        ws["A4"] = name
        ws["B4"] = "2024-W01"
        ws["C4"] = 8
        ws["D4"] = "Week 1"
        ws["E4"] = f"KEY00{idx}"
        ws["F4"] = f"Label {idx}"
        ws["G4"] = f"Function {idx}"
        ws["H4"] = 8.0
        ws["I4"] = "Comment"
        ws["J4"] = "Project A"
        ws["K4"] = "Task 1"

        # Add second row of data
        ws["A5"] = name
        ws["B5"] = "2024-W02"
        ws["C5"] = 7.5
        ws["D5"] = "Week 2"
        ws["E5"] = f"KEY00{idx}"

        ws_lc = wb.create_sheet("LC")
        ws_lc["B3"] = "KEY001"
        ws_lc["C3"] = "Label 1"

        wb.save(rm_folder / f"RM_{name}.xlsx")

    return tmp_path


@pytest.fixture
def template_bytes(tmp_path):
    """
    Create template file bytes for build_interface testing.

    Returns:
        bytes: Binary content of a template Excel file.
    """
    wb = Workbook()
    ws_pointage = wb.active
    ws_pointage.title = "POINTAGE"
    ws_pointage["A2"] = "Week 1"

    ws_lc = wb.create_sheet("LC")
    ws_lc["B3"] = "KEY001"
    ws_lc["C3"] = "Label 1"

    buffer = io.BytesIO()
    wb.save(buffer)

    return buffer.getvalue()
