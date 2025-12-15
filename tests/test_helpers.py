"""
Helper Functions Tests for Roadmap Manager.

Tests for utility functions including XML operations, Excel file manipulation,
data validation, and file system operations.
"""
import xml.etree.ElementTree as ET
import zipfile
from openpyxl import Workbook, load_workbook
from pathlib import Path
import tempfile

from roadmap import __all__ as roadmap_all
from roadmap import RoadmapManager, main as pkg_main
from roadmap import helpers as helpers_module
from roadmap.helpers import (
    add_data_validations_to_sheet,
    build_interface,
    get_collaborators,
    load_lc_excel,
    rmtree_with_retry,
    write_xml,
    zip_folder,
)


class TestGetCollaborators:
    """Tests for get_collaborators function."""

    def test_get_collaborators_from_xml(self, tmp_path):
        """TEST-HELP-001: Verify reading collaborator names from XML file."""
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""

        xml_file = tmp_path / "collabs.xml"
        xml_file.write_text(xml_content, encoding="utf-8")

        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        collabs = get_collaborators(synthese_file)

        assert collabs == ["CLIGNIEZ Yann", "GANI Karim", "MOUHOUT Marouane"]

    def test_get_collaborators_empty_xml(self, tmp_path):
        """TEST-HELP-002: Verify handling of empty collaborators XML."""
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
        </collaborators>"""

        xml_file = tmp_path / "collabs.xml"
        xml_file.write_text(xml_content, encoding="utf-8")

        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        collabs = get_collaborators(synthese_file)

        assert collabs == []

    def test_get_collaborators_missing_xml(self, tmp_path):
        """TEST-HELP-003: Verify handling when XML file doesn't exist."""
        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        collabs = get_collaborators(synthese_file)

        assert collabs == []

    def test_get_collaborators_with_whitespace(self, tmp_path):
        """Verify collaborator names are stripped of whitespace."""
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>  CLIGNIEZ Yann  </collaborator>
            <collaborator>GANI Karim
            </collaborator>
        </collaborators>"""

        xml_file = tmp_path / "collabs.xml"
        xml_file.write_text(xml_content, encoding="utf-8")

        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        collabs = get_collaborators(synthese_file)

        assert collabs == ["CLIGNIEZ Yann", "GANI Karim"]

    def test_get_collaborators_xml_deleted_after_read(self, tmp_path):
        """Verify XML file is deleted after reading."""
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>NAZIH Imane</collaborator>
        </collaborators>"""

        xml_file = tmp_path / "collabs.xml"
        xml_file.write_text(xml_content, encoding="utf-8")

        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        get_collaborators(synthese_file)

        assert not xml_file.exists()

    def test_get_collaborators_warns_when_unlink_fails(self, tmp_path, monkeypatch, caplog):
        """Cover branch where deleting collabs.xml raises an exception."""
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>NAZIH Imane</collaborator>
        </collaborators>"""

        xml_file = tmp_path / "collabs.xml"
        xml_file.write_text(xml_content, encoding="utf-8")
        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        # Patch Path.unlink to fail only for this file
        original_unlink = Path.unlink

        def failing_unlink(path):
            if path == xml_file:
                raise OSError("cannot delete")
            return original_unlink(path)

        monkeypatch.setattr(helpers_module.Path, "unlink", failing_unlink)

        with caplog.at_level("WARNING"):
            collabs = get_collaborators(synthese_file)

        assert collabs == ["NAZIH Imane"]
        # File still exists because unlink failed
        assert xml_file.exists()
        assert "Could not delete XML file" in caplog.text

    def test_get_collaborators_logs_error_on_parse_failure(self, tmp_path, monkeypatch, caplog):
        """Cover outer exception handler when XML parsing fails."""
        # Write invalid XML so that ET.parse raises
        xml_file = tmp_path / "collabs.xml"
        xml_file.write_text("not-xml", encoding="utf-8")
        synthese_file = tmp_path / "Synthèse_RM_CE.xlsm"

        with caplog.at_level("ERROR"):
            collabs = get_collaborators(synthese_file)

        assert collabs == []
        assert "Error reading XML file" in caplog.text


class TestBuildInterface:
    """Tests for build_interface function."""

    def test_build_interface(self, tmp_path, template_bytes):
        """TEST-HELP-004: Verify building collaborator interface from template."""
        output = tmp_path / "RM_YAHYA Oumaima.xlsx"

        build_interface(template_bytes, str(output), "YAHYA Oumaima")

        assert output.exists()

        generated = load_workbook(output)
        assert generated["POINTAGE"]["B1"].value == "YAHYA Oumaima"
        generated.close()

    def test_build_interface_preserves_lc_sheet(self, tmp_path, template_bytes):
        """Verify LC sheet is preserved in generated interface."""
        output = tmp_path / "RM_NAZIH Imane.xlsx"

        build_interface(template_bytes, str(output), "NAZIH Imane")

        generated = load_workbook(output)
        assert "LC" in generated.sheetnames
        generated.close()

    def test_build_interface_adds_data_validations(self, tmp_path, template_bytes):
        """Verify data validations are added to POINTAGE sheet."""
        output = tmp_path / "RM_CLIGNIEZ Yann.xlsx"

        build_interface(template_bytes, str(output), "CLIGNIEZ Yann")

        generated = load_workbook(output)
        ws = generated["POINTAGE"]

        # Check that data validations exist
        assert len(ws.data_validations.dataValidation) == 4
        generated.close()


class TestWriteXml:
    """Tests for write_xml function."""

    def test_write_xml(self, tmp_path):
        """TEST-HELP-005: Verify XML export with row data."""
        rows = [
            ["CLIGNIEZ Yann", "2024-01", 100],
            ["GANI Karim", "2024-01", 200],
        ]

        xml_output = tmp_path / "test_output.xml"
        write_xml(rows, xml_output)

        assert xml_output.exists()

        tree = ET.parse(xml_output)
        root = tree.getroot()

        assert root.tag == "rows"
        assert len(root.findall("row")) == 2

    def test_write_xml_empty(self, tmp_path):
        """TEST-HELP-006: Verify XML export with empty data."""
        rows = []

        xml_output = tmp_path / "test_empty.xml"
        write_xml(rows, xml_output)

        assert xml_output.exists()

        tree = ET.parse(xml_output)
        root = tree.getroot()

        assert root.tag == "rows"
        assert len(root.findall("row")) == 0

    def test_write_xml_with_none(self, tmp_path):
        """TEST-HELP-007: Verify XML export handles None values."""
        rows = [["MOUHOUT Marouane", None, 100]]

        xml_output = tmp_path / "test_none.xml"
        write_xml(rows, xml_output)

        tree = ET.parse(xml_output)
        root = tree.getroot()
        row = root.find("row")

        col2_text = row.find("col2").text
        assert col2_text is None or col2_text == ""

    def test_write_xml_column_names(self, tmp_path):
        """Verify XML uses correct column naming (col1, col2, etc.)."""
        rows = [["A", "B", "C", "D", "E"]]

        xml_output = tmp_path / "test_cols.xml"
        write_xml(rows, xml_output)

        tree = ET.parse(xml_output)
        root = tree.getroot()
        row = root.find("row")

        for i in range(1, 6):
            assert row.find(f"col{i}") is not None

    def test_write_xml_utf8_encoding(self, tmp_path):
        """Verify XML file uses UTF-8 encoding."""
        rows = [["Élève", "Données"]]

        xml_output = tmp_path / "test_utf8.xml"
        write_xml(rows, xml_output)

        content = xml_output.read_text(encoding="utf-8")
        assert 'encoding="utf-8"' in content.lower() or 'encoding=\'utf-8\'' in content.lower()


class TestZipFolder:
    """Tests for zip_folder function."""

    def test_zip_folder(self, tmp_path):
        """TEST-HELP-008: Verify folder compression."""
        source_folder = tmp_path / "source"
        source_folder.mkdir()
        (source_folder / "file1.txt").write_text("content1")
        (source_folder / "file2.txt").write_text("content2")

        zip_path = tmp_path / "archive.zip"
        zip_folder(source_folder, zip_path)

        assert zip_path.exists()

        with zipfile.ZipFile(zip_path, 'r') as zipf:
            names = zipf.namelist()
            assert len(names) == 2

    def test_zip_folder_preserves_structure(self, tmp_path):
        """Verify folder structure is preserved in zip."""
        source_folder = tmp_path / "source"
        source_folder.mkdir()
        subfolder = source_folder / "subfolder"
        subfolder.mkdir()
        (subfolder / "nested.txt").write_text("nested content")

        zip_path = tmp_path / "archive.zip"
        zip_folder(source_folder, zip_path)

        with zipfile.ZipFile(zip_path, 'r') as zipf:
            names = zipf.namelist()
            assert any("subfolder" in name for name in names)

    def test_zip_folder_empty(self, tmp_path):
        """Verify empty folder creates empty zip."""
        source_folder = tmp_path / "empty_source"
        source_folder.mkdir()

        zip_path = tmp_path / "empty_archive.zip"
        zip_folder(source_folder, zip_path)

        assert zip_path.exists()

        with zipfile.ZipFile(zip_path, 'r') as zipf:
            assert len(zipf.namelist()) == 0


class TestRmtreeWithRetry:
    """Tests for rmtree_with_retry function."""

    def test_rmtree_with_retry(self, tmp_path):
        """TEST-HELP-009: Verify folder removal with retry logic."""
        folder = tmp_path / "to_delete"
        folder.mkdir()
        (folder / "file.txt").write_text("content")

        result = rmtree_with_retry(folder)

        assert result is True
        assert not folder.exists()

    def test_rmtree_nested_folders(self, tmp_path):
        """Verify nested folder removal."""
        folder = tmp_path / "parent"
        folder.mkdir()
        child = folder / "child"
        child.mkdir()
        (child / "file.txt").write_text("content")

        result = rmtree_with_retry(folder)

        assert result is True
        assert not folder.exists()

    def test_rmtree_nonexistent_folder(self, tmp_path):
        """Verify handling of non-existent folder."""
        folder = tmp_path / "nonexistent"

        try:
            result = rmtree_with_retry(folder)
        except FileNotFoundError:
            pass

    def test_rmtree_with_retry_zero_retries_returns_false(self, tmp_path):
        """Cover final return False path when max_retries is zero."""
        folder = tmp_path / "to_delete_zero"
        folder.mkdir()

        result = rmtree_with_retry(folder, max_retries=0)

        assert result is False
        assert folder.exists()


class TestAddDataValidations:
    """Tests for add_data_validations_to_sheet function."""

    def test_add_data_validations(self, tmp_path):
        """TEST-HELP-010: Verify data validation lists are added correctly."""
        wb = Workbook()
        ws = wb.active
        ws.title = "POINTAGE"
        wb.create_sheet("LC")

        add_data_validations_to_sheet(ws, start_row=3)

        assert len(ws.data_validations.dataValidation) == 4

        output = tmp_path / "test_validations.xlsx"
        wb.save(output)
        wb.close()

    def test_add_data_validations_clears_existing(self, tmp_path):
        """Verify existing validations are cleared before adding new ones."""
        wb = Workbook()
        ws = wb.active
        ws.title = "POINTAGE"
        wb.create_sheet("LC")

        add_data_validations_to_sheet(ws, start_row=3)
        add_data_validations_to_sheet(ws, start_row=3)

        assert len(ws.data_validations.dataValidation) == 4
        wb.close()

    def test_add_data_validations_creates_list_when_missing_attribute(self):
        """Cover branch where worksheet initially lacks data_validations attribute."""

        class DummySheet:
            def __init__(self):
                self.validations_added = []

            def add_data_validation(self, dv):
                self.validations_added.append(dv)

        sheet = DummySheet()

        add_data_validations_to_sheet(sheet, start_row=3)

        # data_validations attribute should be created even when missing initially
        assert hasattr(sheet, "data_validations")


class TestLoadLcExcel:
    """Tests for load_lc_excel function."""

    def test_load_lc_excel(self, tmp_path):
        """TEST-HELP-011: Verify LC data loading from Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"

        ws["B2"] = "Key1"
        ws["C2"] = "Label1"
        ws["D2"] = "Func1"
        ws["E2"] = "Extra1"
        ws["B3"] = "Key2"
        ws["C3"] = "Label2"
        ws["D3"] = "Func2"

        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        lc_data = load_lc_excel(tmp_path)

        assert len(lc_data) == 2
        assert lc_data[0][0] == "Key1"
        assert lc_data[0][1] == "Label1"

    def test_load_lc_excel_missing_file(self, tmp_path):
        """Verify handling when LC.xlsx doesn't exist."""
        lc_data = load_lc_excel(tmp_path)

        assert lc_data == []

    def test_load_lc_excel_empty_sheet(self, tmp_path):
        """Verify handling of empty LC sheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        lc_data = load_lc_excel(tmp_path)

        assert lc_data == []

    def test_load_lc_excel_deletes_file(self, tmp_path):
        """Verify LC.xlsx is deleted after reading."""
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "Key1"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        load_lc_excel(tmp_path)

        assert not (tmp_path / "LC.xlsx").exists()

    def test_load_lc_excel_missing_lc_sheet(self, tmp_path):
        """Verify handling when LC sheet is missing."""
        wb = Workbook()
        ws = wb.active
        ws.title = "OtherSheet"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        lc_data = load_lc_excel(tmp_path)

        assert lc_data == []

    def test_load_lc_excel_handles_dates_and_text_formatting(self, tmp_path):
        """Cover date-handling branches in load_lc_excel."""
        from datetime import datetime, date

        wb = Workbook()
        ws = wb.active
        ws.title = "LC"

        # Row 2: datetime stored in a text-formatted cell ('@')
        cell_b2 = ws["B2"]
        cell_b2.value = datetime(2024, 1, 2, 15, 30)
        cell_b2.number_format = "@"

        # Row 3: date stored with default date formatting (not '@')
        cell_b3 = ws["B3"]
        cell_b3.value = date(2024, 1, 3)

        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        lc_data = load_lc_excel(tmp_path)

        # Two rows of data should be read
        assert len(lc_data) == 2
        # Both values should be converted to strings (exact representation may vary)
        assert isinstance(lc_data[0][0], str)
        assert isinstance(lc_data[1][0], str)

    def test_load_lc_excel_logs_error_on_invalid_excel(self, tmp_path, caplog):
        """Cover error branch when LC.xlsx is not a valid Excel file."""
        invalid = tmp_path / "LC.xlsx"
        invalid.write_text("not an excel file", encoding="utf-8")

        with caplog.at_level("ERROR"):
            lc_data = load_lc_excel(tmp_path)

        assert lc_data == []
        assert "[LOAD_LC_EXCEL] Error reading Excel file:" in caplog.text

    def test_load_lc_excel_unlink_warning(self, tmp_path, monkeypatch, caplog):
        """Cover warning branch when LC.xlsx cannot be deleted after reading."""
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "Key1"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        excel_file = tmp_path / "LC.xlsx"

        original_unlink = Path.unlink

        def failing_unlink(path):
            if path == excel_file:
                raise OSError("cannot delete")
            return original_unlink(path)

        monkeypatch.setattr(helpers_module.Path, "unlink", failing_unlink)

        with caplog.at_level("WARNING"):
            lc_data = load_lc_excel(tmp_path)

        assert lc_data
        assert excel_file.exists()
        assert "Could not delete Excel file" in caplog.text


class TestGetExeDirAndLoggingPaths:
    """Tests for helpers.get_exe_dir and related logging path logic."""

    def test_get_exe_dir_uses_exe_directory_for_executable(self, monkeypatch, tmp_path):
        """When argv[0] is a .exe, log path should live next to that exe."""
        exe_path = tmp_path / "roadmap.exe"
        exe_path.write_text("")  # ensure path exists

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(exe_path)]
        fake_sys.executable = str(exe_path)
        fake_sys.frozen = False

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        log_path_str = helpers_module.get_exe_dir()
        log_path = Path(log_path_str)

        assert log_path.parent.name == ".logs"
        assert log_path.parent.parent == exe_path.parent

    def test_get_exe_dir_frozen_uses_sys_executable_parent(self, monkeypatch, tmp_path):
        """When running as frozen, use directory of sys.executable."""
        exe_dir = tmp_path / "frozen_dir"
        exe_dir.mkdir()
        fake_exe = exe_dir / "frozen.exe"
        fake_exe.write_text("")

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(fake_exe)]
        fake_sys.executable = str(fake_exe)
        fake_sys.frozen = True

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        log_path = Path(helpers_module.get_exe_dir())
        assert log_path.parent.parent == exe_dir

    def test_get_exe_dir_frozen_non_exe_uses_sys_executable(self, monkeypatch, tmp_path):
        """Cover frozen branch when argv[0] is not an .exe file."""
        exe_dir = tmp_path / "frozen_dir2"
        exe_dir.mkdir()
        fake_exe = exe_dir / "frozen_app.exe"
        fake_exe.write_text("")

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(tmp_path / "launcher")]  # no .exe suffix
        fake_sys.executable = str(fake_exe)
        fake_sys.frozen = True

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        log_path = Path(helpers_module.get_exe_dir())
        assert log_path.parent.parent == exe_dir

    def test_get_exe_dir_no_file_falls_back_to_script_path(self, monkeypatch, tmp_path):
        """Cover NameError branch when __file__ is unavailable."""
        script = tmp_path / "script.py"
        script.write_text("")

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(script)]
        fake_sys.executable = str(script)
        fake_sys.frozen = False

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        g = helpers_module.get_exe_dir.__globals__
        original_file = g.get("__file__")
        g.pop("__file__", None)

        try:
            log_path_str = helpers_module.get_exe_dir()
        finally:
            if original_file is not None:
                g["__file__"] = original_file

        log_path = Path(log_path_str)
        assert log_path.parent.parent == script.parent

    def test_get_exe_dir_fallback_to_cwd_logs_dir(self, monkeypatch, tmp_path):
        """Cover fallback path where creating logs dir next to exe fails once and CWD is used."""
        exe_dir = tmp_path / "exe_dir"
        exe_dir.mkdir()
        script = exe_dir / "launcher.exe"
        script.write_text("")

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(script)]
        fake_sys.executable = str(script)
        fake_sys.frozen = False

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        original_mkdir = helpers_module.Path.mkdir
        call_count = {"n": 0}

        def fake_mkdir(self, *args, **kwargs):
            call_count["n"] += 1
            if call_count["n"] == 1:
                raise OSError("fail exe logs")
            return original_mkdir(self, *args, **kwargs)

        monkeypatch.setattr(helpers_module.Path, "mkdir", fake_mkdir)

        log_path_str = helpers_module.get_exe_dir()
        log_path = Path(log_path_str)

        assert call_count["n"] >= 2
        assert log_path.exists() or log_path.parent.exists()

    def test_get_exe_dir_fallback_to_temp_logs_dir(self, monkeypatch, tmp_path):
        """Cover final fallback to temp directory when both exe and CWD logs dirs fail."""
        exe_dir = tmp_path / "exe_dir2"
        exe_dir.mkdir()
        script = exe_dir / "launcher2.exe"
        script.write_text("")

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(script)]
        fake_sys.executable = str(script)
        fake_sys.frozen = False

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        original_mkdir = helpers_module.Path.mkdir
        call_count = {"n": 0}

        def fake_mkdir(self, *args, **kwargs):
            call_count["n"] += 1
            if call_count["n"] <= 2:
                raise OSError("fail logs")
            return original_mkdir(self, *args, **kwargs)

        monkeypatch.setattr(helpers_module.Path, "mkdir", fake_mkdir)

        log_path_str = helpers_module.get_exe_dir()
        log_path = Path(log_path_str)

        assert Path(tempfile.gettempdir()) in log_path.parents

    def test_get_exe_dir_raises_oserror_when_logs_dir_not_created(self, monkeypatch, tmp_path):
        """Cover branch that raises OSError when logs_dir.exists() returns False after mkdir."""
        exe_dir = tmp_path / "exe_dir_check"
        exe_dir.mkdir()
        script = exe_dir / "launcher_check.exe"
        script.write_text("")

        fake_sys = type("FakeSys", (), {})()
        fake_sys.argv = [str(script)]
        fake_sys.executable = str(script)
        fake_sys.frozen = False

        monkeypatch.setattr(helpers_module, "sys", fake_sys)

        original_mkdir = helpers_module.Path.mkdir
        original_exists = helpers_module.Path.exists

        def fake_mkdir(self, *args, **kwargs):
            # Let mkdir succeed normally
            return original_mkdir(self, *args, **kwargs)

        def fake_exists(self):
            # For the first logs_dir check, pretend it does not exist to trigger the raise
            if self.name == ".logs" and self.parent == exe_dir:
                return False
            return original_exists(self)

        monkeypatch.setattr(helpers_module.Path, "mkdir", fake_mkdir)
        monkeypatch.setattr(helpers_module.Path, "exists", fake_exists)

        # get_exe_dir should still succeed by falling back to CWD after the internal OSError
        log_path_str = helpers_module.get_exe_dir()
        assert log_path_str.endswith("roadmap.log")


class TestRmtreeWithRetryEdgeCases:
    """Additional edge‑case tests for rmtree_with_retry."""

    def test_rmtree_with_retry_fails_after_retries(self, monkeypatch, tmp_path):
        """Simulate persistent PermissionError to cover retry and failure branch."""
        folder = tmp_path / "locked"
        folder.mkdir()

        def always_fail(path, onerror=None):
            raise PermissionError("locked")

        monkeypatch.setattr(helpers_module.shutil, "rmtree", always_fail)

        result = rmtree_with_retry(folder, max_retries=2)

        assert result is False

    def test_rmtree_on_rm_error_calls_func(self, monkeypatch, tmp_path):
        """Cover on_rm_error path where func(path) is executed without errors."""
        folder = tmp_path / "with_readonly"
        folder.mkdir()

        called = {"count": 0}

        def fake_rmtree(path, onerror=None):
            # Simulate rmtree invoking onerror callback, which should call our func(path)
            onerror(lambda p: called.__setitem__("count", called["count"] + 1), path, None)

        monkeypatch.setattr(helpers_module.shutil, "rmtree", fake_rmtree)

        result = rmtree_with_retry(folder, max_retries=1)

        assert result is True
        assert called["count"] == 1


class TestRoadmapPackageInit:
    """Light‑weight checks for roadmap.__init__ exports."""

    def test_package_exports(self):
        """Ensure __all__ lists the public API and they are importable."""
        assert "RoadmapManager" in roadmap_all
        assert "main" in roadmap_all

        # Accessing imported symbols should work
        assert RoadmapManager is not None
        assert callable(pkg_main)
