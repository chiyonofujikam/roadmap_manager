"""
RoadmapManager Integration Tests for Roadmap Manager.

Tests for the RoadmapManager class including initialization,
interface creation, deletion, pointage export, and LC updates.
"""
import xml.etree.ElementTree as ET
from pathlib import Path
import shutil
import tempfile

from openpyxl import Workbook, load_workbook

import roadmap.roadmap as roadmap_module
from roadmap.roadmap import RoadmapManager


# Module-level function for pickling in multiprocessing tests
def _failing_build_interface(template_bytes, output_path, collab):
    """Mock build_interface that raises an exception for testing."""
    raise RuntimeError("build failed")


class TestRoadmapManagerInit:
    """Tests for RoadmapManager initialization."""

    def test_init_with_valid_files(self, setup_test_environment):
        """TEST-INT-001: Verify RoadmapManager initializes with correct paths."""
        tmp_path = setup_test_environment

        manager = RoadmapManager(tmp_path)

        assert manager.base_path == tmp_path
        assert manager.synthese_file == tmp_path / "Synthèse_RM_CE.xlsm"
        assert manager.template_file == tmp_path / "RM_template.xlsx"
        assert manager.rm_folder == tmp_path / "RM_Collaborateurs"
        assert manager.archived_folder == tmp_path / "Archived"
        assert manager.deleted_folder == tmp_path / "Deleted"
        assert manager.all_ok is True

    def test_init_missing_files(self, tmp_path):
        """TEST-INT-002: Verify RoadmapManager handles missing files."""
        manager = RoadmapManager(tmp_path)

        assert manager.all_ok is False

    def test_init_missing_template(self, tmp_path):
        """Verify handling when only synthese file exists."""
        (tmp_path / "Synthèse_RM_CE.xlsm").touch()

        manager = RoadmapManager(tmp_path)

        assert manager.all_ok is False

    def test_init_missing_synthese(self, tmp_path):
        """Verify handling when only template file exists."""
        (tmp_path / "RM_template.xlsx").touch()

        manager = RoadmapManager(tmp_path)

        assert manager.all_ok is False

    def test_init_creates_folders(self, tmp_path):
        """Verify initialization creates required folders."""
        (tmp_path / "Synthèse_RM_CE.xlsm").touch()
        (tmp_path / "RM_template.xlsx").touch()

        manager = RoadmapManager(tmp_path)

        assert (tmp_path / "RM_Collaborateurs").exists()
        assert (tmp_path / "Archived").exists()
        assert (tmp_path / "Deleted").exists()


class TestCreateInterfaces:
    """Tests for interface creation methods."""

    def test_create_interfaces_normal(self, setup_test_environment):
        """TEST-INT-003: Verify interface creation in normal mode."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        manager.create_interfaces()

        rm_folder = tmp_path / "RM_Collaborateurs"
        created_files = list(rm_folder.glob("RM_*.xlsx"))

        assert len(created_files) == 3

        # Verify file names match collaborators
        file_names = {f.name for f in created_files}
        assert "RM_CLIGNIEZ Yann.xlsx" in file_names
        assert "RM_GANI Karim.xlsx" in file_names
        assert "RM_MOUHOUT Marouane.xlsx" in file_names

    def test_create_interfaces_parallel(self, setup_test_environment):
        """TEST-INT-004: Verify interface creation in parallel mode."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        manager.create_interfaces_fast(max_workers=2)

        rm_folder = tmp_path / "RM_Collaborateurs"
        created_files = list(rm_folder.glob("RM_*.xlsx"))

        assert len(created_files) == 3

    def test_create_interfaces_skip_existing(self, setup_test_environment):
        """TEST-INT-005: Verify existing files are not overwritten."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        rm_folder = tmp_path / "RM_Collaborateurs"
        existing_file = rm_folder / "RM_CLIGNIEZ Yann.xlsx"

        # Create existing file with marker
        wb = Workbook()
        ws = wb.active
        ws.title = "POINTAGE"
        ws["A1"] = "EXISTING_MARKER"
        ws["B1"] = "CLIGNIEZ Yann"
        wb.create_sheet("LC")
        wb.save(existing_file)
        wb.close()

        manager.create_interfaces()

        # Verify marker still exists
        wb = load_workbook(existing_file)
        assert wb["POINTAGE"]["A1"].value == "EXISTING_MARKER"
        wb.close()

    def test_create_interfaces_empty_collaborators(self, tmp_path):
        """Verify handling when no collaborators in list."""
        # Create required files
        (tmp_path / "Synthèse_RM_CE.xlsm").touch()

        template = Workbook()
        template.active.title = "POINTAGE"
        template.create_sheet("LC")
        template.save(tmp_path / "RM_template.xlsx")
        template.close()

        # Create empty collabs.xml
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager = RoadmapManager(tmp_path)
        manager.create_interfaces()

        rm_folder = tmp_path / "RM_Collaborateurs"
        created_files = list(rm_folder.glob("RM_*.xlsx"))

        assert len(created_files) == 0

    def test_create_interfaces_all_exist(self, setup_test_environment_with_interfaces):
        """Verify no action when all files already exist."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Get file modification times before
        rm_folder = tmp_path / "RM_Collaborateurs"
        before_times = {f.name: f.stat().st_mtime for f in rm_folder.glob("RM_*.xlsx")}

        # Need to recreate collabs.xml since it was deleted
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager.create_interfaces()

        # Files should not be modified (times unchanged)
        after_times = {f.name: f.stat().st_mtime for f in rm_folder.glob("RM_*.xlsx")}
        assert before_times == after_times

    def test_created_interface_has_collaborator_name(self, setup_test_environment):
        """Verify created interface has collaborator name in B1."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        manager.create_interfaces()

        rm_folder = tmp_path / "RM_Collaborateurs"
        collab_file = rm_folder / "RM_CLIGNIEZ Yann.xlsx"

        wb = load_workbook(collab_file)
        assert wb["POINTAGE"]["B1"].value == "CLIGNIEZ Yann"
        wb.close()

    def test_created_interface_has_data_validations(self, setup_test_environment):
        """Verify created interface has data validations."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        manager.create_interfaces()

        rm_folder = tmp_path / "RM_Collaborateurs"
        collab_file = rm_folder / "RM_CLIGNIEZ Yann.xlsx"

        wb = load_workbook(collab_file)
        ws = wb["POINTAGE"]

        assert len(ws.data_validations.dataValidation) == 4
        wb.close()

    def test_create_interfaces_fast_template_permission_error(self, setup_test_environment, monkeypatch):
        """Cover PermissionError branch when reading template bytes in fast mode."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        original_read_bytes = Path.read_bytes

        def fake_read_bytes(path: Path, *args, **kwargs):
            if path == manager.template_file:
                raise PermissionError("template locked")
            return original_read_bytes(path, *args, **kwargs)

        monkeypatch.setattr(roadmap_module.Path, "read_bytes", fake_read_bytes)

        # Should return early without raising
        manager.create_interfaces_fast(max_workers=1)

    def test_create_interfaces_template_permission_error(self, setup_test_environment, monkeypatch):
        """Cover PermissionError branch in sequential create_interfaces when template is locked."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        original_load = roadmap_module.load_workbook

        def fake_load(path, *args, **kwargs):
            if Path(path) == manager.template_file:
                raise PermissionError("locked")
            return original_load(path, *args, **kwargs)

        monkeypatch.setattr(roadmap_module, "load_workbook", fake_load)

        # Should log an error and return without raising
        manager.create_interfaces()


class TestDeleteInterfaces:
    """Tests for interface deletion methods."""

    def test_delete_and_archive(self, setup_test_environment_with_interfaces):
        """TEST-INT-006: Verify deletion with archiving."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        manager.delete_and_archive_interfaces(archive=True)

        rm_folder = tmp_path / "RM_Collaborateurs"
        archived_folder = tmp_path / "Archived"
        deleted_folder = tmp_path / "Deleted"

        # RM folder should be deleted or empty
        assert not rm_folder.exists() or len(list(rm_folder.glob("*.xlsx"))) == 0

        # Archive should have a zip file
        assert len(list(archived_folder.glob("*.zip"))) > 0

        # Deleted should have a zip file
        assert len(list(deleted_folder.glob("*.zip"))) > 0

    def test_delete_without_archive(self, setup_test_environment_with_interfaces):
        """TEST-INT-007: Verify deletion without archiving."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        manager.delete_and_archive_interfaces(archive=False)

        rm_folder = tmp_path / "RM_Collaborateurs"
        archived_folder = tmp_path / "Archived"
        deleted_folder = tmp_path / "Deleted"

        # RM folder should be deleted or empty
        assert not rm_folder.exists() or len(list(rm_folder.glob("*.xlsx"))) == 0

        # Archive should be empty
        assert len(list(archived_folder.glob("*.zip"))) == 0

        # Deleted should have a zip file
        assert len(list(deleted_folder.glob("*.zip"))) > 0

    def test_delete_empty_folder(self, setup_test_environment):
        """Verify handling when RM_Collaborateurs is empty."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        # RM_Collaborateurs exists but is empty
        manager.delete_and_archive_interfaces(archive=True)

        # Should complete without error
        rm_folder = tmp_path / "RM_Collaborateurs"
        assert not rm_folder.exists()

    def test_delete_nonexistent_folder(self, tmp_path):
        """Verify handling when RM_Collaborateurs doesn't exist."""
        (tmp_path / "Synthèse_RM_CE.xlsm").touch()
        (tmp_path / "RM_template.xlsx").touch()

        manager = RoadmapManager(tmp_path)

        # Should complete without error
        manager.delete_and_archive_interfaces(archive=True)

    def test_delete_and_archive_handles_archive_error(self, setup_test_environment_with_interfaces, monkeypatch):
        """Cover error branch when archiving RM_Collaborateurs fails."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        def fail_zip(folder, dest):
            raise RuntimeError("zip failed")

        monkeypatch.setattr(roadmap_module, "zip_folder", fail_zip)

        # With archive=True an error while archiving should cause early return
        manager.delete_and_archive_interfaces(archive=True)

        # RM folder should still exist because delete phase was not reached
        assert manager.rm_folder.exists()

    def test_delete_and_archive_handles_deleted_zip_error(self, setup_test_environment_with_interfaces, monkeypatch):
        """Cover error branch when creating Deleted zip fails."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        original_zip = roadmap_module.zip_folder

        def selective_zip(src, dest):
            # Let archive phase succeed, fail only for Deleted_RM_Collaborateurs_*.zip
            if "Deleted_RM_Collaborateurs_" in dest.name:
                raise RuntimeError("deleted zip error")
            return original_zip(src, dest)

        monkeypatch.setattr(roadmap_module, "zip_folder", selective_zip)

        manager.delete_and_archive_interfaces(archive=True)

        # RM folder should still exist because error happened during deleted-zip phase
        assert manager.rm_folder.exists()


class TestDeleteMissingCollaborators:
    """Tests for cleanup of missing collaborators."""

    def test_delete_missing_collaborators(self, setup_test_environment_with_interfaces):
        """TEST-INT-008: Verify cleanup of orphaned interface files."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        rm_folder = tmp_path / "RM_Collaborateurs"

        # Create orphan file
        orphan_file = rm_folder / "RM_Orphan User.xlsx"
        wb = Workbook()
        wb.active.title = "POINTAGE"
        wb.save(orphan_file)
        wb.close()

        # Create new collabs.xml without Orphan
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager.delete_missing_collaborators()

        assert not orphan_file.exists()
        # Other files should still exist
        assert (rm_folder / "RM_CLIGNIEZ Yann.xlsx").exists()

    def test_delete_missing_no_orphans(self, setup_test_environment_with_interfaces):
        """Verify no action when all files match collaborators."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        rm_folder = tmp_path / "RM_Collaborateurs"
        before_count = len(list(rm_folder.glob("RM_*.xlsx")))

        # Create matching collabs.xml
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager.delete_missing_collaborators()

        after_count = len(list(rm_folder.glob("RM_*.xlsx")))
        assert before_count == after_count

    def test_delete_missing_creates_archive(self, setup_test_environment_with_interfaces):
        """Verify orphan files are archived before deletion."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        rm_folder = tmp_path / "RM_Collaborateurs"
        deleted_folder = tmp_path / "Deleted"

        # Create orphan file
        orphan_file = rm_folder / "RM_Orphan User.xlsx"
        wb = Workbook()
        wb.active.title = "POINTAGE"
        wb.save(orphan_file)
        wb.close()

        # Create new collabs.xml without Orphan
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager.delete_missing_collaborators()

        # Check archive was created
        archives = list(deleted_folder.glob("Deleted_Missing_*.zip"))
        assert len(archives) > 0

    def test_delete_missing_collaborators_folder_missing(self, setup_test_environment_with_interfaces, caplog):
        """Cover branch where RM_Collaborateurs folder does not exist."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Remove folder before calling
        shutil.rmtree(manager.rm_folder)

        with caplog.at_level("WARNING"):
            manager.delete_missing_collaborators()

        assert "[DELETE_MISSING_COLLABORATORS] RM_Collaborateurs folder does not exist" in caplog.text

    def test_delete_missing_collaborators_no_collaborators(self, setup_test_environment_with_interfaces, caplog):
        """Cover branch where collabs.xml exists but has no collaborators."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        with caplog.at_level("WARNING"):
            manager.delete_missing_collaborators()

        assert "[DELETE_MISSING_COLLABORATORS] No collaborators found in XML. Skipping cleanup." in caplog.text

    def test_delete_missing_collaborators_no_existing_files(self, setup_test_environment, caplog):
        """Cover branch where RM_Collaborateurs has no .xlsx files."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        rm_folder = manager.rm_folder
        # Ensure folder exists but remove any files
        for f in rm_folder.glob("*.xlsx"):
            f.unlink()

        # Create collabs.xml with collaborators so get_collaborators is non-empty
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        with caplog.at_level("INFO"):
            manager.delete_missing_collaborators()

        assert "[DELETE_MISSING_COLLABORATORS] No files found in RM_Collaborateurs folder" in caplog.text

    def test_delete_missing_collaborators_zip_error_and_cleanup_warning(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover zip-creation error and temp-folder cleanup warning branches."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        rm_folder = manager.rm_folder
        deleted_folder = manager.deleted_folder

        # Create orphan files so there is something to delete
        orphan1 = rm_folder / "RM_Orphan1.xlsx"
        orphan2 = rm_folder / "RM_Orphan2.xlsx"
        wb = Workbook()
        wb.active.title = "POINTAGE"
        wb.save(orphan1)
        wb.save(orphan2)
        wb.close()

        # collabs.xml lists no orphans, so orphans become files_to_delete
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        # Force zip_folder to fail when creating Deleted_Missing_*.zip
        original_zip = roadmap_module.zip_folder

        def failing_zip(src, dest):
            if "Deleted_Missing_RM_collaborators_" in dest.name:
                raise RuntimeError("zip error")
            return original_zip(src, dest)

        monkeypatch.setattr(roadmap_module, "zip_folder", failing_zip)

        # Also make shutil.rmtree for temp folder fail to hit cleanup warning
        original_rmtree = shutil.rmtree

        def failing_rmtree(path, *args, **kwargs):
            if "missing_collabs_" in str(path):
                raise RuntimeError("rmtree error")
            return original_rmtree(path, *args, **kwargs)

        monkeypatch.setattr(roadmap_module.shutil, "rmtree", failing_rmtree)

        with caplog.at_level("ERROR"):
            manager.delete_missing_collaborators()

        # Even though archive creation failed, deletion should have proceeded
        assert not orphan1.exists() or not orphan2.exists()
        assert "Error creating zip archive" in caplog.text

    def test_delete_missing_collaborators_unlink_permission_and_error(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover PermissionError and generic Exception branches during unlink."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        rm_folder = manager.rm_folder

        orphan1 = rm_folder / "RM_OrphanPerm.xlsx"
        orphan2 = rm_folder / "RM_OrphanErr.xlsx"
        wb = Workbook()
        wb.active.title = "POINTAGE"
        wb.save(orphan1)
        wb.save(orphan2)
        wb.close()

        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        # Patch Path.unlink to raise different exceptions per orphan
        original_unlink = Path.unlink

        def custom_unlink(path):
            if path == orphan1:
                raise PermissionError("locked")
            if path == orphan2:
                raise RuntimeError("generic error")
            return original_unlink(path)

        monkeypatch.setattr(roadmap_module.Path, "unlink", custom_unlink)

        with caplog.at_level("WARNING"):
            manager.delete_missing_collaborators()

        assert "[DELETE_MISSING_COLLABORATORS] Cannot delete RM_OrphanPerm.xlsx - file may be open in Excel" in caplog.text
        # Error for orphan2 should be logged at ERROR level
        assert "[DELETE_MISSING_COLLABORATORS] Error deleting RM_OrphanErr.xlsx: generic error" in caplog.text


class TestPointage:
    """Tests for pointage export functionality."""

    def test_pointage_export(self, setup_test_environment_with_data):
        """TEST-INT-009: Verify pointage data export to XML."""
        tmp_path = setup_test_environment_with_data
        manager = RoadmapManager(tmp_path)

        result = manager.pointage()

        assert result is True

        xml_output = tmp_path / "pointage_output.xml"
        assert xml_output.exists()

        tree = ET.parse(xml_output)
        root = tree.getroot()
        rows = root.findall("row")

        # Should have 6 rows (2 per collaborator × 3 collaborators)
        assert len(rows) == 6

    def test_pointage_empty_interfaces(self, setup_test_environment_with_interfaces):
        """TEST-INT-010: Verify pointage export with empty interfaces."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        result = manager.pointage()

        assert result is False

        xml_output = tmp_path / "pointage_output.xml"
        assert xml_output.exists()

        tree = ET.parse(xml_output)
        root = tree.getroot()
        assert len(root.findall("row")) == 0

    def test_pointage_no_interfaces(self, setup_test_environment):
        """Verify pointage with no interface files."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        result = manager.pointage()

        assert result is False

    def test_pointage_skips_temp_files(self, setup_test_environment_with_data):
        """Verify pointage skips temporary Excel files (~$)."""
        tmp_path = setup_test_environment_with_data
        manager = RoadmapManager(tmp_path)

        rm_folder = tmp_path / "RM_Collaborateurs"

        # Create a temp file
        temp_file = rm_folder / "~$RM_CLIGNIEZ Yann.xlsx"
        temp_file.write_text("temp")

        result = manager.pointage()

        # Should still work
        assert result is True

    def test_pointage_rm_folder_missing_logs_error(self, setup_test_environment_with_data, caplog):
        """Cover branch where RM_Collaborateurs folder is missing."""
        tmp_path = setup_test_environment_with_data
        manager = RoadmapManager(tmp_path)

        # Remove the RM_Collaborateurs folder after setup
        shutil.rmtree(manager.rm_folder)

        with caplog.at_level("ERROR"):
            result = manager.pointage()

        assert result is False
        assert "RM_Collaborateurs folder not found" in caplog.text


class TestUpdateLc:
    """Tests for LC update functionality."""

    def test_update_lc(self, setup_test_environment_with_interfaces):
        """TEST-INT-011: Verify LC sheet update in all files."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Create LC.xlsx with test data
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "NewKey"
        ws["C2"] = "NewLabel"
        ws["D2"] = "NewFunc"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        manager.update_lc()

        # Verify template updated
        template_wb = load_workbook(tmp_path / "RM_template.xlsx")
        assert template_wb["LC"]["B2"].value == "NewKey"
        template_wb.close()

        # Verify interface files updated
        rm_folder = tmp_path / "RM_Collaborateurs"
        collab_wb = load_workbook(rm_folder / "RM_CLIGNIEZ Yann.xlsx")
        assert collab_wb["LC"]["B2"].value == "NewKey"
        collab_wb.close()

    def test_update_lc_no_file(self, setup_test_environment_with_interfaces):
        """Verify handling when LC.xlsx doesn't exist."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Should complete without error
        manager.update_lc()

    def test_update_lc_no_interfaces(self, setup_test_environment):
        """Verify LC update works with only template."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        # Create LC.xlsx with test data
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "NewKey"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        manager.update_lc()

        # Verify template updated
        template_wb = load_workbook(tmp_path / "RM_template.xlsx")
        assert template_wb["LC"]["B2"].value == "NewKey"
        template_wb.close()

    def test_update_lc_recreates_data_validations(self, setup_test_environment_with_interfaces):
        """Verify data validations are recreated after LC update."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Create LC.xlsx with test data
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "NewKey"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        manager.update_lc()

        # Verify interface files have data validations
        rm_folder = tmp_path / "RM_Collaborateurs"
        collab_wb = load_workbook(rm_folder / "RM_CLIGNIEZ Yann.xlsx")
        ws = collab_wb["POINTAGE"]

        assert len(ws.data_validations.dataValidation) == 4
        collab_wb.close()

    def test_update_lc_all_ok_false_logs_error(self, tmp_path, caplog):
        """Cover branch where required files are missing (all_ok is False)."""
        manager = RoadmapManager(tmp_path)

        assert manager.all_ok is False

        with caplog.at_level("ERROR"):
            manager.update_lc()

        assert "[UPDATE_LC] Required files are missing. Cannot proceed." in caplog.text

    def test_update_lc_template_update_exception(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover template update error branch inside update_lc."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Prepare LC.xlsx with simple data
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "Key1"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        original_update = manager._update_lc_in_file

        def failing_update(file_path, lc_data):
            if Path(file_path) == manager.template_file:
                raise RuntimeError("template error")
            return original_update(file_path, lc_data)

        manager._update_lc_in_file = failing_update

        with caplog.at_level("ERROR"):
            manager.update_lc()

        assert "[UPDATE_LC] Error updating template file: template error" in caplog.text

    def test_update_lc_rm_file_update_exception(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover collaborator file update error branch inside update_lc."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Prepare LC.xlsx with simple data
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "Key1"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        original_update = manager._update_lc_in_file

        def failing_update(file_path, lc_data):
            # Let template update succeed; fail for first RM_ file
            if Path(file_path).name.startswith("RM_"):
                raise RuntimeError("rm update error")
            return original_update(file_path, lc_data)

        manager._update_lc_in_file = failing_update

        with caplog.at_level("ERROR"):
            manager.update_lc()

        assert "[UPDATE_LC] Error updating RM_CLIGNIEZ Yann.xlsx: rm update error" in caplog.text

    def test_update_lc_no_data_logs_warning(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover branch where LC.xlsx exists but load_lc_excel returns no data."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Create LC.xlsx but force loader to return empty list
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        monkeypatch.setattr(roadmap_module, "load_lc_excel", lambda base_dir: [])

        with caplog.at_level("WARNING"):
            manager.update_lc()

        assert "[UPDATE_LC] No LC data found in LC.xlsx. Nothing to update." in caplog.text

    def test_update_lc_in_file_handles_copy_permission_error(self, setup_test_environment_with_interfaces, monkeypatch):
        """Cover PermissionError branch in _update_lc_in_file copy step."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        target_file = manager.template_file

        def fake_copy2(src, dst, *args, **kwargs):
            if Path(src) == target_file:
                raise PermissionError("locked")
            return shutil.copy2(src, dst, *args, **kwargs)

        monkeypatch.setattr(roadmap_module.shutil, "copy2", fake_copy2)

        # Should not raise even if copy fails
        manager._update_lc_in_file(target_file, [["Key", "Label"]])

    def test_update_lc_in_file_missing_lc_sheet(self, tmp_path):
        """Cover branch where LC sheet is missing from workbook."""
        # Create simple workbook without LC sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Other"
        excel_path = tmp_path / "no_lc.xlsx"
        wb.save(excel_path)
        wb.close()

        manager = RoadmapManager(tmp_path)
        # Mark all_ok to True so helper can run; template/synthese not required here
        manager.all_ok = True

        # Should not raise even though LC sheet is missing
        manager._update_lc_in_file(excel_path, [["Key", "Label"]])

    def test_update_lc_in_file_generic_exception_logs_error(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover generic exception handler inside _update_lc_in_file."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        target_file = manager.template_file

        # Force load_workbook on temp_path to raise a generic exception
        original_load = roadmap_module.load_workbook

        def failing_load(path, *args, **kwargs):
            # _update_lc_in_file loads from the temp copy, not from target_file itself.
            # We treat any path not equal to target_file as the temp one and raise.
            if Path(path) != target_file:
                raise RuntimeError("load error")
            return original_load(path, *args, **kwargs)

        monkeypatch.setattr(roadmap_module, "load_workbook", failing_load)

        with caplog.at_level("ERROR"):
            manager._update_lc_in_file(target_file, [["Key", "Label"]])

        assert f"[UPDATE_LC] Error updating LC in {target_file.name}: load error" in caplog.text

    def test_update_lc_in_file_tempfile_unlink_warning(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover warning branch when temporary file cannot be deleted."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Patch Path.unlink used inside roadmap_module to always fail for temp files
        original_unlink = roadmap_module.Path.unlink

        def failing_unlink(path):
            # Fail only for files in the system temp directory to avoid impacting test data
            tmpdir = Path(tempfile.gettempdir())
            if tmpdir in Path(path).parents:
                raise PermissionError("cannot delete temp")
            return original_unlink(path)

        monkeypatch.setattr(roadmap_module.Path, "unlink", failing_unlink)

        with caplog.at_level("WARNING"):
            manager._update_lc_in_file(manager.template_file, [["Key", "Label"]])

        assert "Could not delete temporary file" in caplog.text


class TestEdgeCases:
    """Tests for edge cases and error handling."""

    def test_all_ok_false_prevents_operations(self, tmp_path):
        """Verify operations are skipped when all_ok is False."""
        manager = RoadmapManager(tmp_path)

        assert manager.all_ok is False

        # These should return early without error
        manager.create_interfaces()
        manager.create_interfaces_fast()
        manager.delete_and_archive_interfaces(archive=True)
        manager.delete_missing_collaborators()
        result = manager.pointage()

        assert result is False

    def test_path_as_string(self, setup_test_environment):
        """Verify manager accepts path as string."""
        tmp_path = setup_test_environment

        manager = RoadmapManager(str(tmp_path))

        assert manager.all_ok is True
        assert manager.base_path == tmp_path

    def test_special_characters_in_collaborator_name(self, tmp_path):
        """Verify handling of special characters in collaborator names."""
        # Create required files
        template = Workbook()
        template.active.title = "POINTAGE"
        template.create_sheet("LC")
        template.save(tmp_path / "RM_template.xlsx")
        template.close()

        (tmp_path / "Synthèse_RM_CE.xlsm").touch()

        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>NAZIH Imane</collaborator>
            <collaborator>YAHYA Oumaima</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager = RoadmapManager(tmp_path)
        manager.create_interfaces()

        rm_folder = tmp_path / "RM_Collaborateurs"
        assert (rm_folder / "RM_NAZIH Imane.xlsx").exists()
        assert (rm_folder / "RM_YAHYA Oumaima.xlsx").exists()


class TestCreateInterfacesFastCoverage:
    """Additional tests for create_interfaces_fast() to cover missing lines."""

    def test_create_interfaces_fast_empty_collaborators(self, tmp_path, caplog):
        """Cover empty collaborators list branch in create_interfaces_fast()."""
        # Create required files
        (tmp_path / "Synthèse_RM_CE.xlsm").touch()

        template = Workbook()
        template.active.title = "POINTAGE"
        template.create_sheet("LC")
        template.save(tmp_path / "RM_template.xlsx")
        template.close()

        # Create empty collabs.xml
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        manager = RoadmapManager(tmp_path)
        with caplog.at_level("INFO"):
            manager.create_interfaces_fast()

        assert "[CREATE_INTERFACES] the list of CE is empty" in caplog.text

    def test_create_interfaces_fast_file_already_exists_debug(self, setup_test_environment_with_interfaces, caplog):
        """Cover debug log when file already exists in create_interfaces_fast()."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Recreate collabs.xml since it was deleted
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        # Files already exist from fixture
        with caplog.at_level("DEBUG"):
            manager.create_interfaces_fast(max_workers=1)

        assert "[CREATE_INTERFACES] File already exists:" in caplog.text

    def test_create_interfaces_fast_all_files_exist(self, setup_test_environment_with_interfaces, caplog):
        """Cover branch when all files already exist in create_interfaces_fast()."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Recreate collabs.xml since it was deleted
        xml_content = """<?xml version="1.0" encoding="UTF-8"?>
        <collaborators>
            <collaborator>CLIGNIEZ Yann</collaborator>
            <collaborator>GANI Karim</collaborator>
            <collaborator>MOUHOUT Marouane</collaborator>
        </collaborators>"""
        (tmp_path / "collabs.xml").write_text(xml_content, encoding="utf-8")

        with caplog.at_level("INFO"):
            manager.create_interfaces_fast(max_workers=1)

        assert "[CREATE_INTERFACES] All collaborator files already exist. Nothing to create." in caplog.text

    def test_create_interfaces_fast_exception_in_parallel(self, setup_test_environment, monkeypatch, caplog):
        """Cover exception handling branch in parallel processing."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        # Mock build_interface in both helpers and roadmap modules
        # ProcessPoolExecutor needs the function to be picklable and the same object
        import roadmap.helpers as helpers_module
        
        monkeypatch.setattr(helpers_module, "build_interface", _failing_build_interface)
        monkeypatch.setattr(roadmap_module, "build_interface", _failing_build_interface)

        with caplog.at_level("ERROR"):
            manager.create_interfaces_fast(max_workers=1)

        assert "error: build failed" in caplog.text


class TestDeleteInterfacesCoverage:
    """Additional tests for delete_and_archive_interfaces() to cover missing lines."""

    def test_delete_nonexistent_folder_warning(self, tmp_path, caplog):
        """Cover warning branch when RM_Collaborateurs doesn't exist."""
        (tmp_path / "Synthèse_RM_CE.xlsm").touch()
        (tmp_path / "RM_template.xlsx").touch()

        manager = RoadmapManager(tmp_path)
        # Remove the folder if it was created during init
        if manager.rm_folder.exists():
            shutil.rmtree(manager.rm_folder)

        with caplog.at_level("WARNING"):
            manager.delete_and_archive_interfaces(archive=False)

        assert "[DELETE_INTERFACES] RM_Collaborateurs folder does not exist" in caplog.text

    def test_delete_empty_folder_exception(self, setup_test_environment, monkeypatch, caplog):
        """Cover exception branch when removing empty folder fails."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        # Mock rmtree_with_retry to raise an exception
        def failing_rmtree(path):
            raise RuntimeError("rmtree failed")

        monkeypatch.setattr(roadmap_module, "rmtree_with_retry", failing_rmtree)

        with caplog.at_level("ERROR"):
            manager.delete_and_archive_interfaces(archive=False)

        assert "[DELETE_INTERFACES] Error removing empty folder: rmtree failed" in caplog.text

    def test_delete_rmtree_returns_false_warning(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover warning branch when rmtree_with_retry returns False."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Mock rmtree_with_retry to return False
        def false_rmtree(path):
            return False

        monkeypatch.setattr(roadmap_module, "rmtree_with_retry", false_rmtree)

        with caplog.at_level("WARNING"):
            manager.delete_and_archive_interfaces(archive=False)

        assert "[DELETE_INTERFACES] Could not remove original folder, but zip was created" in caplog.text


class TestPointageCoverage:
    """Additional tests for pointage() to cover missing lines."""

    def test_pointage_empty_row_break(self, setup_test_environment):
        """Cover break branch when hitting empty row in pointage()."""
        tmp_path = setup_test_environment
        manager = RoadmapManager(tmp_path)

        rm_folder = tmp_path / "RM_Collaborateurs"

        # Create interface file with data followed by explicitly empty row
        wb = Workbook()
        ws = wb.active
        ws.title = "POINTAGE"
        ws["B1"] = "CLIGNIEZ Yann"
        ws["K1"] = 8.0

        # Add data row (row 4)
        ws["A4"] = "CLIGNIEZ Yann"
        ws["B4"] = "2024-W01"
        ws["C4"] = 8
        ws["D4"] = "Week 1"
        ws["E4"] = "KEY001"
        ws["F4"] = "Label 1"
        ws["G4"] = "Function 1"
        ws["H4"] = 8.0
        ws["I4"] = "Comment"
        ws["J4"] = "Project A"
        ws["K4"] = "Task 1"

        # Explicitly set row 5 to None values to ensure it's detected as empty
        # This ensures the break statement at line 450 is executed
        for col in range(1, 12):  # Columns A-K (1-11)
            cell = ws.cell(row=5, column=col)
            cell.value = None

        wb.create_sheet("LC")
        wb.save(rm_folder / "RM_CLIGNIEZ Yann.xlsx")
        wb.close()

        result = manager.pointage()

        # Should process only one row (row 4) and break at row 5
        assert result is True
        xml_output = tmp_path / "pointage_output.xml"
        assert xml_output.exists()
        tree = ET.parse(xml_output)
        root = tree.getroot()
        rows = root.findall("row")
        # Should have exactly 1 row (row 4), row 5 should trigger break
        assert len(rows) == 1


class TestUpdateLcCoverage:
    """Additional tests for update_lc() to cover missing lines."""

    def test_update_lc_copy_generic_exception(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover generic exception branch when copying file in update_lc()."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Create LC.xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "NewKey"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        # Mock shutil.copy2 to raise a generic exception (not PermissionError)
        original_copy2 = roadmap_module.shutil.copy2

        def failing_copy2(src, dst, *args, **kwargs):
            # Only fail for collaborator files, not template
            if Path(src).name.startswith("RM_"):
                raise RuntimeError("copy failed")
            return original_copy2(src, dst, *args, **kwargs)

        monkeypatch.setattr(roadmap_module.shutil, "copy2", failing_copy2)

        with caplog.at_level("WARNING"):
            manager.update_lc()

        assert "[UPDATE_LC] Error copying" in caplog.text

    def test_update_lc_data_validation_exception(self, setup_test_environment_with_interfaces, monkeypatch, caplog):
        """Cover exception branch when recreating data validation fails."""
        tmp_path = setup_test_environment_with_interfaces
        manager = RoadmapManager(tmp_path)

        # Create LC.xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "LC"
        ws["B2"] = "NewKey"
        wb.save(tmp_path / "LC.xlsx")
        wb.close()

        # Mock add_data_validations_to_sheet to raise an exception
        def failing_add_validations(ws, start_row):
            raise RuntimeError("validation failed")

        monkeypatch.setattr(roadmap_module, "add_data_validations_to_sheet", failing_add_validations)

        with caplog.at_level("ERROR"):
            manager.update_lc()

        assert "[UPDATE_LC] Error recreating data validation" in caplog.text